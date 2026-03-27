"""Main evaluation pipeline: load -> 3x evaluate -> Claude judge -> log -> export."""

import asyncio
import json
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl

from tutor_eval.config import (
    CRITERIA,
    CRITERIA_NAMES_RU,
    DATA_DIR,
    EVAL_TEMPERATURE,
    EVALUATOR_MODELS,
    LOGS_DIR,
    RESULTS_DIR,
    RUBRICS_PATH,
    EVALUATOR_PROMPT_PATH,
)
from tutor_eval.judge import run_judge
from tutor_eval.loader import Dialog, load_all, load_xlsx
from tutor_eval.math_check import check_dialog_math, result_to_dict as math_result_to_dict
from tutor_eval.providers.base import EvalResult
from tutor_eval.providers.claude_provider import ClaudeJudgeProvider
from tutor_eval.providers.gemini_provider import OpenRouterEvalProvider


# ---------------------------------------------------------------------------
# Incremental JSONL writer
# ---------------------------------------------------------------------------

class JsonlWriter:
    """Append-mode JSONL writer — flushes after every record."""

    def __init__(self, path: Path):
        self.path = path
        self._fh = open(path, "a", encoding="utf-8")

    def write(self, record: dict):
        self._fh.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
        self._fh.flush()

    def close(self):
        self._fh.close()


# ---------------------------------------------------------------------------
# Prompt building
# ---------------------------------------------------------------------------

def build_eval_prompt(dialog: Dialog) -> str:
    """Build evaluator prompt for a single dialog."""
    rubrics = RUBRICS_PATH.read_text(encoding="utf-8")
    template = EVALUATOR_PROMPT_PATH.read_text(encoding="utf-8")
    return (
        template
        .replace("{rubrics}", rubrics)
        .replace("{task}", dialog.task)
        .replace("{grade_group}", dialog.grade_group)
        .replace("{theme}", f"{dialog.theme} \u2192 {dialog.subtheme}")
        .replace("{dialog}", dialog.text)
    )


# ---------------------------------------------------------------------------
# Single-dialog evaluation
# ---------------------------------------------------------------------------

async def evaluate_dialog(
    dialog: Dialog,
    evaluators: list[OpenRouterEvalProvider],
    judge: ClaudeJudgeProvider,
    w_raw: JsonlWriter,
    w_judge: JsonlWriter,
    w_errors: JsonlWriter,
) -> dict | None:
    """Evaluate a single dialog with 3 evaluator runs, then Claude judge."""
    prompt = build_eval_prompt(dialog)

    tasks = [_safe_evaluate(ev, prompt, dialog.dialog_id, w_errors) for ev in evaluators]
    results: list[EvalResult | None] = await asyncio.gather(*tasks)
    valid_results = [r for r in results if r is not None]

    for r in valid_results:
        w_raw.write({
            "dialog_id": dialog.dialog_id,
            "evaluator": r.evaluator,
            "model": r.model,
            "temperature": r.temperature,
            "criteria": r.scores,
            "critical_flags": r.critical_flags,
            "overall_notes": r.overall_notes,
            "tokens_used": r.tokens_used,
            "latency_ms": r.latency_ms,
            "timestamp": datetime.now().isoformat(),
        })

    if len(valid_results) < 2:
        w_errors.write({
            "dialog_id": dialog.dialog_id,
            "error": f"Only {len(valid_results)} evaluators succeeded, need at least 2",
            "timestamp": datetime.now().isoformat(),
        })
        return None

    try:
        decision = await run_judge(dialog, valid_results, judge)
    except Exception as e:
        w_errors.write({
            "dialog_id": dialog.dialog_id,
            "error": f"Judge failed: {e}",
            "timestamp": datetime.now().isoformat(),
        })
        return None

    w_judge.write({
        "dialog_id": dialog.dialog_id,
        "final_scores": decision.get("final_scores", {}),
        "agreement": decision.get("agreement", {}),
        "overrides": decision.get("overrides", []),
        "critical_flags": decision.get("critical_flags", {}),
        "_meta": decision.get("_meta", {}),
        "timestamp": datetime.now().isoformat(),
    })

    return decision


async def _safe_evaluate(evaluator, prompt: str, dialog_id: str, w_errors: JsonlWriter) -> EvalResult | None:
    try:
        return await evaluator.evaluate(prompt)
    except Exception as e:
        w_errors.write({
            "dialog_id": dialog_id,
            "evaluator": evaluator.name,
            "error": str(e),
            "timestamp": datetime.now().isoformat(),
        })
        return None


# ---------------------------------------------------------------------------
# Resume support
# ---------------------------------------------------------------------------

def _load_completed_ids(log_dir: Path, result_dir: Path) -> set[str]:
    """Load dialog_ids that were already evaluated in a previous run."""
    completed = set()
    for parent in [result_dir, log_dir]:
        for fname in ["all_judge_decisions.jsonl", "judge_decisions.jsonl"]:
            p = parent / fname
            if p.exists():
                with open(p, encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            try:
                                record = json.loads(line)
                                completed.add(record["dialog_id"])
                            except (json.JSONDecodeError, KeyError):
                                pass
    return completed


# ---------------------------------------------------------------------------
# Pipeline orchestration
# ---------------------------------------------------------------------------

async def run_pipeline(
    dialogs: list[Dialog],
    evaluators: list[OpenRouterEvalProvider] | None = None,
    judge: ClaudeJudgeProvider | None = None,
    concurrency: int = 5,
    resume_run: str | None = None,
    skip_math: bool = False,
) -> Path:
    """Run the full evaluation pipeline.

    Args:
        resume_run: timestamp of a previous run to resume (e.g. "2026-03-25_14-54")
        skip_math: skip math correctness check stage
    Returns:
        path to the results directory.
    """
    if evaluators is None:
        evaluators = [
            OpenRouterEvalProvider(model=m["model"], name=m["name"], temperature=EVAL_TEMPERATURE)
            for m in EVALUATOR_MODELS
        ]
    if judge is None:
        judge = ClaudeJudgeProvider()

    # Use existing run dirs for resume, or create new ones
    if resume_run:
        ts = resume_run
    else:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M")

    log_dir = LOGS_DIR / ts
    result_dir = RESULTS_DIR / ts
    log_dir.mkdir(parents=True, exist_ok=True)
    result_dir.mkdir(parents=True, exist_ok=True)

    # Skip already-evaluated dialogs on resume
    skip_ids = set()
    if resume_run:
        skip_ids = _load_completed_ids(log_dir, result_dir)
        if skip_ids:
            print(f"Resume: found {len(skip_ids)} already evaluated, skipping them")

    remaining = [d for d in dialogs if d.dialog_id not in skip_ids]

    eval_names = [e.name for e in evaluators]
    print(f"Pipeline: {len(remaining)} dialogs ({len(dialogs)} total, {len(skip_ids)} skipped)")
    print(f"Evaluators: {eval_names}, judge: Claude")
    print(f"Logs: {log_dir}")
    print(f"Results: {result_dir}")
    print()

    # Write dialog metadata upfront (so dashboard can show texts during run)
    dialog_meta = []
    for d in dialogs:
        dialog_meta.append({
            "dialog_id": d.dialog_id,
            "text": d.text,
            "task": d.task,
            "file_name": d.file_name,
            "student_type": d.student_type,
            "student_model": d.student_model,
            "grade_group": d.grade_group,
            "theme": d.theme,
            "subtheme": d.subtheme,
        })
    _write_jsonl(result_dir / "dialogs.jsonl", dialog_meta)

    if not remaining:
        print("Nothing to do — all dialogs already evaluated.")
        return result_dir

    # Open incremental writers (append mode for resume)
    w_raw = JsonlWriter(log_dir / "raw_scores.jsonl")
    w_judge = JsonlWriter(log_dir / "judge_decisions.jsonl")
    w_errors = JsonlWriter(log_dir / "errors.jsonl")

    all_decisions = {}
    semaphore = asyncio.Semaphore(concurrency)
    completed = 0
    errors = 0

    async def process_one(dialog: Dialog):
        nonlocal completed, errors
        async with semaphore:
            decision = await evaluate_dialog(
                dialog, evaluators, judge, w_raw, w_judge, w_errors,
            )
            if decision is not None:
                all_decisions[dialog.dialog_id] = decision
            else:
                errors += 1

            completed += 1
            if completed % 10 == 0 or completed == len(remaining):
                print(f"  Progress: {completed}/{len(remaining)} dialogs ({errors} errors)")

    await asyncio.gather(*[process_one(d) for d in remaining])

    w_raw.close()
    w_judge.close()
    w_errors.close()

    # Rebuild full results: load previous decisions + new ones
    all_judge_records = []
    all_raw_records = []
    full_decisions = {}

    # Load previous records if resuming
    for fname in ["judge_decisions.jsonl", "all_judge_decisions.jsonl"]:
        p = log_dir / fname
        if p.exists():
            with open(p, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line:
                        record = json.loads(line)
                        all_judge_records.append(record)
                        full_decisions[record["dialog_id"]] = record

    for fname in ["raw_scores.jsonl", "all_raw_scores.jsonl"]:
        p = log_dir / fname
        if p.exists():
            with open(p, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line:
                        all_raw_records.append(json.loads(line))

    # Deduplicate (keep latest per dialog_id for judge)
    seen_judge = set()
    deduped_judge = []
    for r in reversed(all_judge_records):
        if r["dialog_id"] not in seen_judge:
            seen_judge.add(r["dialog_id"])
            deduped_judge.append(r)
    deduped_judge.reverse()

    # Write consolidated results
    _write_jsonl(result_dir / "all_raw_scores.jsonl", all_raw_records)
    _write_jsonl(result_dir / "all_judge_decisions.jsonl", deduped_judge)

    # Write summary and xlsx using full decisions
    summary = _build_summary(dialogs, full_decisions, all_raw_records, errors)
    (log_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_results_xlsx(result_dir / "scores.xlsx", dialogs, full_decisions)

    total_evaluated = len(full_decisions)
    print(f"\nDone! {total_evaluated}/{len(dialogs)} dialogs evaluated.")
    print(f"New in this run: {len(all_decisions)}, errors: {errors}")
    print(f"Results: {result_dir / 'scores.xlsx'}")

    # --- Math correctness check ---
    if not skip_math:
        print(f"\n--- Math correctness check ---")
        math_results = await _run_math_check(dialogs, result_dir, concurrency)
        if math_results:
            total_claims = sum(r.get("claims_count", 0) for r in math_results)
            total_incorrect = sum(r.get("incorrect_count", 0) for r in math_results)
            dialogs_with_err = sum(1 for r in math_results if r.get("incorrect_count", 0) > 0)
            error_rate = dialogs_with_err / len(math_results) if math_results else 0
            print(f"Math check: {total_claims} claims, {total_incorrect} incorrect")
            print(f"Error rate: {error_rate:.1%} ({dialogs_with_err}/{len(math_results)} dialogs)")

    return result_dir


# ---------------------------------------------------------------------------
# Math correctness integration
# ---------------------------------------------------------------------------

async def _run_math_check(
    dialogs: list[Dialog],
    result_dir: Path,
    concurrency: int,
) -> list[dict]:
    """Run math correctness check on all dialogs and save to result_dir."""
    math_model = "google/gemini-3-flash-preview"
    extractor = OpenRouterEvalProvider(model=math_model, name="math-extractor", temperature=0.0)
    codegen = OpenRouterEvalProvider(model=math_model, name="math-codegen", temperature=0.0)

    sem = asyncio.Semaphore(concurrency)
    total = len(dialogs)
    done = 0
    math_errors = 0

    async def process_one(dialog: Dialog) -> dict | None:
        nonlocal done, math_errors
        async with sem:
            try:
                result = await check_dialog_math(dialog, extractor, codegen)
                done += 1
                if result.incorrect_count > 0:
                    print(f"  Math [{done}/{total}] {dialog.dialog_id}: {result.incorrect_count} INCORRECT")
                elif done % 10 == 0 or done == total:
                    print(f"  Math [{done}/{total}] progress...")
                return math_result_to_dict(result)
            except Exception as e:
                done += 1
                math_errors += 1
                print(f"  Math [{done}/{total}] ERROR {dialog.dialog_id}: {e}")
                return {
                    "dialog_id": dialog.dialog_id,
                    "error": str(e),
                    "claims_count": 0, "correct_count": 0,
                    "incorrect_count": 0, "error_count": 0,
                    "verifications": [],
                }

    print(f"Math check: {total} dialogs, model={math_model}")
    results = await asyncio.gather(*[process_one(d) for d in dialogs])
    results = [r for r in results if r]

    out_path = result_dir / "math_check_results.jsonl"
    _write_jsonl(out_path, results)
    print(f"Math results: {out_path}")

    return results


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_jsonl(path: Path, records: list[dict]):
    with open(path, "w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")


def _build_summary(
    dialogs: list[Dialog],
    decisions: dict,
    raw_scores: list,
    error_count: int,
) -> dict:
    criterion_scores = {c: [] for c in CRITERIA}
    agreement_counts = {"unanimous": 0, "majority": 0, "split": 0}
    override_count = 0

    for dialog_id, decision in decisions.items():
        if isinstance(decision, str):
            continue
        final = decision.get("final_scores", {})
        agreement = decision.get("agreement", {})
        overrides = decision.get("overrides", [])

        for c in CRITERIA:
            entry = final.get(c, {})
            score = entry.get("score")
            if score is not None and int(score) >= 0:
                criterion_scores[c].append(int(score))

            level = agreement.get(c, "")
            if level in agreement_counts:
                agreement_counts[level] += 1

        override_count += len(overrides)

    avg_scores = {}
    for c in CRITERIA:
        scores = criterion_scores[c]
        avg_scores[c] = {
            "mean": round(sum(scores) / len(scores), 2) if scores else 0,
            "count": len(scores),
        }

    total_tokens = sum(r.get("tokens_used", 0) for r in raw_scores)
    for d in decisions.values():
        if isinstance(d, dict):
            total_tokens += d.get("_meta", {}).get("tokens_used", 0)

    total_agreement = sum(agreement_counts.values()) or 1
    return {
        "total_dialogs": len(dialogs),
        "evaluated": len(decisions),
        "errors": error_count,
        "avg_scores": avg_scores,
        "agreement": {
            k: round(v / total_agreement * 100, 1)
            for k, v in agreement_counts.items()
        },
        "override_count": override_count,
        "total_tokens": total_tokens,
        "pipeline": {
            "evaluators": [m["model"] for m in EVALUATOR_MODELS],
            "temperature": EVAL_TEMPERATURE,
            "judge": "anthropic/claude-sonnet-4",
        },
    }


def _write_results_xlsx(path: Path, dialogs: list[Dialog], decisions: dict):
    wb = openpyxl.Workbook()

    # Sheet 1: Per-dialog scores
    ws = wb.active
    ws.title = "По диалогам"

    headers = [
        "dialog_id", "файл", "тип_ученика", "модель_ученика",
        "задача", "класс", "тема",
    ] + [CRITERIA_NAMES_RU[c] for c in CRITERIA] + ["средний", "уровень_согласия", "криты"]

    ws.append(headers)

    dialog_map = {d.dialog_id: d for d in dialogs}

    for dialog_id, decision in sorted(decisions.items()):
        if isinstance(decision, str):
            continue
        d = dialog_map.get(dialog_id)
        if not d:
            continue

        final = decision.get("final_scores", {})
        agreement = decision.get("agreement", {})

        scores = []
        for c in CRITERIA:
            s = final.get(c, {}).get("score")
            scores.append(int(s) if s is not None else "")

        numeric_scores = [s for s in scores if isinstance(s, int) and s >= 0]
        avg = round(sum(numeric_scores) / len(numeric_scores), 2) if numeric_scores else ""

        levels = [agreement.get(c, "") for c in CRITERIA]
        if all(l == "unanimous" for l in levels):
            overall_agreement = "unanimous"
        elif "split" in levels:
            overall_agreement = "mixed (has splits)"
        else:
            overall_agreement = "mostly majority"

        # Crits
        crits = decision.get("critical_flags", {})
        confirmed = crits.get("confirmed", {})
        crit_parts = []
        for crit_name, crit_info in confirmed.items():
            cat = crit_info.get("category") if isinstance(crit_info, dict) else None
            crit_parts.append(f"{crit_name}:{cat}" if cat else crit_name)
        crit_str = ", ".join(crit_parts)

        row = [
            d.dialog_id, d.file_name, d.student_type, d.student_model,
            d.task[:100], d.grade_group, d.theme,
        ] + scores + [avg, overall_agreement, crit_str]

        ws.append(row)

    # Sheet 2: Per-file aggregation
    ws2 = wb.create_sheet("По файлам")
    ws2.append(["файл", "тип_ученика", "модель_ученика", "кол-во"] +
               [CRITERIA_NAMES_RU[c] for c in CRITERIA] + ["средний"])

    file_groups: dict[str, list] = {}
    for dialog_id, decision in decisions.items():
        if isinstance(decision, str):
            continue
        d = dialog_map.get(dialog_id)
        if d:
            file_groups.setdefault(d.file_name, []).append(decision)

    for fname, decs in sorted(file_groups.items()):
        d_sample = next(
            (dialog_map[did] for did in decisions
             if dialog_map.get(did) and dialog_map[did].file_name == fname),
            None,
        )

        per_criterion = {c: [] for c in CRITERIA}
        for dec in decs:
            final = dec.get("final_scores", {})
            for c in CRITERIA:
                s = final.get(c, {}).get("score")
                if s is not None and int(s) >= 0:
                    per_criterion[c].append(int(s))

        avgs = []
        all_scores = []
        for c in CRITERIA:
            vals = per_criterion[c]
            avg = round(sum(vals) / len(vals), 2) if vals else ""
            avgs.append(avg)
            if isinstance(avg, float):
                all_scores.append(avg)

        total_avg = round(sum(all_scores) / len(all_scores), 2) if all_scores else ""

        ws2.append([
            fname,
            d_sample.student_type if d_sample else "",
            d_sample.student_model if d_sample else "",
            len(decs),
        ] + avgs + [total_avg])

    # Sheet 3: Per-criterion
    ws3 = wb.create_sheet("По критериям")
    ws3.append(["критерий", "средний балл", "кол-во оценок", "мин", "макс"])

    for c in CRITERIA:
        all_vals = []
        for decision in decisions.values():
            if isinstance(decision, str):
                continue
            s = decision.get("final_scores", {}).get(c, {}).get("score")
            if s is not None and int(s) >= 0:
                all_vals.append(int(s))

        ws3.append([
            CRITERIA_NAMES_RU[c],
            round(sum(all_vals) / len(all_vals), 2) if all_vals else "",
            len(all_vals),
            min(all_vals) if all_vals else "",
            max(all_vals) if all_vals else "",
        ])

    wb.save(path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

async def main():
    """CLI: python -m tutor_eval.pipeline [--sample N] [--file PATH] [--concurrency N] [--resume RUN_ID]"""
    import argparse

    parser = argparse.ArgumentParser(description="AI Tutor Evaluation Pipeline")
    parser.add_argument("--sample", type=int, default=0, help="Evaluate only N random dialogs (0 = all)")
    parser.add_argument("--file", type=str, default="", help="Evaluate a single xlsx file")
    parser.add_argument("--data-dir", type=str, default="", help="Directory with xlsx files")
    parser.add_argument("--concurrency", type=int, default=5, help="Max concurrent evaluations")
    parser.add_argument("--resume", type=str, default="", help="Resume a previous run by timestamp (e.g. 2026-03-25_14-54)")
    parser.add_argument("--filter-model", type=str, default="", help="Only evaluate dialogs whose student_model contains this substring (e.g. gemini3flash)")
    parser.add_argument("--skip-math", action="store_true", help="Skip math correctness check")
    args = parser.parse_args()

    if args.file:
        print(f"Loading: {args.file}")
        dialogs = load_xlsx(args.file)
    elif args.data_dir:
        print(f"Loading all from: {args.data_dir}")
        dialogs = load_all(args.data_dir)
    else:
        print(f"Loading all from: {DATA_DIR}")
        dialogs = load_all(DATA_DIR)

    if args.filter_model:
        before = len(dialogs)
        dialogs = [d for d in dialogs if args.filter_model in d.student_model]
        print(f"Filtered by model '{args.filter_model}': {before} -> {len(dialogs)} dialogs")

    if args.sample > 0:
        import random
        random.seed(42)
        dialogs = random.sample(dialogs, min(args.sample, len(dialogs)))
        print(f"Sampled {len(dialogs)} dialogs")

    evaluators = [
        OpenRouterEvalProvider(model=m["model"], name=m["name"], temperature=EVAL_TEMPERATURE)
        for m in EVALUATOR_MODELS
    ]
    print(f"Evaluators: {[e.name for e in evaluators]}")

    result_dir = await run_pipeline(
        dialogs=dialogs,
        evaluators=evaluators,
        concurrency=args.concurrency,
        resume_run=args.resume or None,
        skip_math=args.skip_math,
    )


if __name__ == "__main__":
    asyncio.run(main())
