"""Анализ качества интентов: match rate, распределение, проблемные реплики."""
from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

st.set_page_config(page_title="Качество интентов", page_icon="🎯", layout="wide")

DATA_DIR = Path(__file__).parent.parent / "data" / "intent_eval"
RESULTS_PATH = DATA_DIR / "results.csv"

# ── Палитры и константы ──────────────────────────────────────────────────────

COLOR_MATCH = "#2ecc71"
COLOR_MISMATCH = "#e74c3c"
COLOR_WARN = "#f39c12"

MODEL_COLORS = {
    "DeepSeek-V3.1": "#8e44ad",
    "Gemini-2.5-Flash": "#e67e22",
    "Gemini-3-Flash": "#3498db",
    "GLM-4.5": "#2ecc71",
}

STUDENT_LABELS = {
    "weak": "Слабый",
    "medium": "Средний",
    "strong": "Сильный",
}
STUDENT_ORDER = ["strong", "medium", "weak"]
STUDENT_LABEL_ORDER = ["Сильный", "Средний", "Слабый"]

DEFAULT_WEIGHTS = {
    "answer": 58,
    "set-problem": 14,
    "agree-with-tutor": 12,
    "chat": 5,
    "get-explanation": 4,
    "criticize-tutor": 3,
    "get-solution": 2,
    "find-mistake": 1,
    "end-dialog": 1,
    "thank-tutor": 0,
}

AUDIT_INTENTS = [
    {"id": "answer (correct)", "name": "Верный ответ", "default_weight": "58% × P(верно)", "prompt_file": "intent_answer_correct.md"},
    {"id": "answer (careless)", "name": "Невнимательность", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "answer (procedure)", "name": "Незнание процедуры", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "answer (method)", "name": "Незнание способа", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "answer (unstable_proc)", "name": "Неустойч. процедура", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "answer (unstable_method)", "name": "Неустойч. способ", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "answer (misconception)", "name": "Неверная концепция", "default_weight": "58% × P(ош) × W", "prompt_file": "intent_answer_wrong.md"},
    {"id": "set-problem", "name": "Задать задачу", "default_weight": "14%", "prompt_file": "intent_set_problem.md"},
    {"id": "agree-with-tutor", "name": "Согласиться", "default_weight": "12%", "prompt_file": "intent_agree_with_tutor.md"},
    {"id": "chat", "name": "Болтовня", "default_weight": "5%", "prompt_file": "intent_chat.md"},
    {"id": "get-explanation", "name": "Просить объяснение", "default_weight": "4%", "prompt_file": "intent_get_explanation.md"},
    {"id": "criticize-tutor", "name": "Критиковать", "default_weight": "3%", "prompt_file": "intent_criticize_tutor.md"},
    {"id": "get-solution", "name": "Попросить ответ", "default_weight": "2%", "prompt_file": "intent_get_solution.md"},
    {"id": "find-mistake", "name": "Найти ошибку", "default_weight": "1%", "prompt_file": "intent_find_mistake.md"},
    {"id": "end-dialog", "name": "Закончить", "default_weight": "1%", "prompt_file": "intent_end_dialog.md"},
    {"id": "thank-tutor", "name": "Поблагодарить", "default_weight": "0%", "prompt_file": "intent_thank_tutor.md"},
]


# ── Загрузка данных ──────────────────────────────────────────────────────────


OVERRIDES_PATH = DATA_DIR / "human_overrides.json"


def _mtime(p: Path) -> float:
    return p.stat().st_mtime if p.exists() else 0.0


def _turn_key(row) -> str:
    return f"{row.get('model','')}|{row.get('student_type','')}|{row.get('dialog_idx','')}|{row.get('turn_idx','')}"


def load_overrides() -> dict:
    if not OVERRIDES_PATH.exists():
        return {}
    with open(OVERRIDES_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_overrides(overrides: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(OVERRIDES_PATH, "w", encoding="utf-8") as f:
        json.dump(overrides, f, ensure_ascii=False, indent=2)


@st.cache_data
def load_results(_mtime_key: float, _overrides_mtime: float) -> pd.DataFrame | None:
    if not RESULTS_PATH.exists():
        return None
    df = pd.read_csv(RESULTS_PATH)
    if "match" in df.columns:
        df["match"] = df["match"].astype(str).str.lower().map(
            {"true": True, "1": True, "1.0": True, "false": False, "0": False, "0.0": False}
        )
    if "confidence" in df.columns:
        df["confidence"] = pd.to_numeric(df["confidence"], errors="coerce")
    if "model" not in df.columns:
        df["model"] = "unknown"
    if "student_type" not in df.columns:
        df["student_type"] = "unknown"
    df["model"] = df["model"].fillna("unknown")
    df["student_type"] = df["student_type"].fillna("unknown")

    _bool_map = {"true": True, "1": True, "1.0": True, "false": False, "0": False, "0.0": False}
    if "is_defect" in df.columns:
        df["is_defect"] = df["is_defect"].astype(str).str.lower().map(_bool_map).fillna(False)
    if "ctx_appropriate" in df.columns:
        df["ctx_appropriate"] = df["ctx_appropriate"].astype(str).str.lower().map(_bool_map).fillna(True)

    # Apply human overrides
    df["human_override"] = False
    df["human_reason"] = ""
    df["reviewed"] = False
    overrides = load_overrides()
    if overrides:
        for idx, row in df.iterrows():
            key = _turn_key(row)
            if key in overrides:
                ov = overrides[key]
                if "is_defect" in ov and ov["is_defect"] is not None and "is_defect" in df.columns:
                    df.at[idx, "is_defect"] = ov["is_defect"]
                    df.at[idx, "human_override"] = True
                if "match" in ov and ov["match"] is not None:
                    df.at[idx, "match"] = ov["match"]
                    df.at[idx, "human_override"] = True
                if "ctx_appropriate" in ov and ov["ctx_appropriate"] is not None and "ctx_appropriate" in df.columns:
                    df.at[idx, "ctx_appropriate"] = ov["ctx_appropriate"]
                if "actual_intent" in ov and ov["actual_intent"]:
                    df.at[idx, "actual_intent"] = ov["actual_intent"]
                if ov.get("human_reason"):
                    df.at[idx, "human_override"] = True
                df.at[idx, "human_reason"] = ov.get("human_reason", "")
                df.at[idx, "reviewed"] = ov.get("reviewed", False)

    return df


def _parse_dialog_to_turns(dialog_text: str) -> list[dict]:
    """Parse dialog text into structured turns."""
    turns: list[dict] = []
    lines = dialog_text.strip().split("\n")
    current_role = None
    current_intent = None
    current_text_lines: list[str] = []

    def flush():
        if current_role and current_text_lines:
            turns.append({
                "role": current_role,
                "intent": current_intent,
                "text": "\n".join(current_text_lines).strip(),
            })

    for line in lines:
        m = re.match(r"^Пользователь\s*\[([^\]]+)\]\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "user"
            current_intent = m.group(1).strip()
            current_text_lines = [m.group(2).strip()] if m.group(2).strip() else []
            continue
        m = re.match(r"^Пользователь\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "user"
            current_intent = None
            current_text_lines = [m.group(1).strip()] if m.group(1).strip() else []
            continue
        m = re.match(r"^Ассистент\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "assistant"
            current_intent = None
            current_text_lines = [m.group(1).strip()] if m.group(1).strip() else []
            continue
        if current_role:
            current_text_lines.append(line)

    flush()
    return turns


SOURCE_DIALOGS_PATH = DATA_DIR / "source_dialogs.xlsx"


@st.cache_data
def load_source_dialogs(_mtime_key: float) -> pd.DataFrame | None:
    """Load source dialogs from source_dialogs.xlsx and expand by model×student_type from results."""
    if not SOURCE_DIALOGS_PATH.exists():
        return None
    import openpyxl
    wb = openpyxl.load_workbook(SOURCE_DIALOGS_PATH, data_only=True)
    ws = wb.active
    base_records = []
    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if len(row) < 19:
            continue
        dialog_text = row[18]
        if not dialog_text or not str(dialog_text).strip():
            continue
        if str(dialog_text).strip() in ("string", "any", "int64", "float64"):
            continue
        base_records.append({
            "dialog_idx": row_idx - 2,
            "grade_group": str(row[2] or ""),
            "task_id": str(row[8] or ""),
            "dialog": str(dialog_text),
        })
    wb.close()
    if not base_records:
        return None
    # Expand each dialog by model×student_type combinations present in results
    results_df = pd.read_csv(RESULTS_PATH)
    combos = results_df[["model", "student_type"]].drop_duplicates().to_dict("records")
    all_records = []
    for rec in base_records:
        for combo in combos:
            all_records.append({**rec, **combo})
    return pd.DataFrame(all_records)


def _match_rate_color(rate: float) -> str:
    if rate >= 0.8:
        return COLOR_MATCH
    elif rate >= 0.6:
        return COLOR_WARN
    return COLOR_MISMATCH


# ── Заголовок ────────────────────────────────────────────────────────────────

st.title("🎯 Качество интентов")
st.caption("Анализ совпадения объявленных и фактических интентов в диалогах")

df_all = load_results(_mtime(RESULTS_PATH), _mtime(OVERRIDES_PATH))

if df_all is None:
    st.warning(
        "Файл с результатами не найден.\n\n"
        f"**Ожидаемый путь:** `{RESULTS_PATH}`\n\n"
        "Запустите пайплайн оценки интентов и положите `results.csv` в указанную папку."
    )
    st.stop()

# ── Сайдбар: фильтры ────────────────────────────────────────────────────────

with st.sidebar:
    st.header("Фильтры")

    available_models = sorted(df_all["model"].unique().tolist())
    selected_models = st.multiselect(
        "Модель",
        options=available_models,
        default=available_models,
        key="filter_model",
    )

    available_students = [s for s in STUDENT_ORDER if s in df_all["student_type"].values]
    selected_students = st.multiselect(
        "Тип ученика",
        options=available_students,
        default=available_students,
        format_func=lambda x: STUDENT_LABELS.get(x, x),
        key="filter_student",
    )

    st.markdown("---")
    st.caption(f"Всего записей: {len(df_all):,}")

# Apply filters
df = df_all[
    df_all["model"].isin(selected_models) & df_all["student_type"].isin(selected_students)
].copy()

if df.empty:
    st.warning("Нет данных по выбранным фильтрам.")
    st.stop()

st.sidebar.caption(f"После фильтров: {len(df):,}")

# ── Табы ─────────────────────────────────────────────────────────────────────

tab_overview, tab_models, tab_context, tab_review, tab_distrib, tab_problems, tab_answer, tab_dialogs, tab_audit, tab_reference = st.tabs(
    ["Обзор", "Модели", "Контекст", "Проверка", "Распределение", "Проблемные реплики", "Answer детально", "Диалоги", "Аудит", "Справка"]
)

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 1: Обзор
# ═══════════════════════════════════════════════════════════════════════════════

with tab_overview:
    total_turns = len(df)
    overall_match = df["match"].mean() if "match" in df.columns else 0.0
    n_models = df["model"].nunique()
    n_students = df["student_type"].nunique()

    has_ctx = "ctx_appropriate" in df.columns
    ctx_rate = df["ctx_appropriate"].mean() if has_ctx else None

    has_defect = "is_defect" in df.columns
    n_defects = int(df["is_defect"].sum()) if has_defect else 0

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Реплик", f"{total_turns:,}")
    c2.metric("Брак", f"{n_defects:,} ({100*n_defects/total_turns:.1f}%)" if total_turns else "0")
    c3.metric("Intent match", f"{overall_match:.1%}")
    c4.metric("Context appropriate", f"{ctx_rate:.1%}" if ctx_rate is not None else "N/A")
    c5.metric("Моделей", n_models)
    c6.metric("Типов ученика", n_students)

    st.markdown("---")

    # ── Match rate по интентам (горизонтальные бары) ──
    st.subheader("Match rate по интентам")

    intent_stats = (
        df.groupby("intent_declared")
        .agg(n_turns=("match", "size"), match_rate=("match", "mean"))
        .reset_index()
        .sort_values("match_rate", ascending=True)
    )

    intent_stats["color"] = intent_stats["match_rate"].apply(_match_rate_color)
    intent_stats["match_pct"] = (intent_stats["match_rate"] * 100).round(1)

    fig_bar = go.Figure()
    fig_bar.add_trace(
        go.Bar(
            y=intent_stats["intent_declared"],
            x=intent_stats["match_pct"],
            orientation="h",
            marker_color=intent_stats["color"],
            text=intent_stats["match_pct"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>Match rate: %{x:.1f}%<br><extra></extra>",
        )
    )
    fig_bar.update_layout(
        xaxis_title="Match rate, %",
        yaxis_title="",
        height=max(400, len(intent_stats) * 35),
        margin=dict(l=10, r=40, t=10, b=40),
        xaxis=dict(range=[0, 105]),
    )
    fig_bar.add_vline(x=80, line_dash="dash", line_color=COLOR_MATCH, annotation_text="80%", annotation_position="top")
    fig_bar.add_vline(x=60, line_dash="dash", line_color=COLOR_WARN, annotation_text="60%", annotation_position="top")
    st.plotly_chart(fig_bar, use_container_width=True)

    # ── Match rate по типам учеников (горизонтальные бары) ──
    st.subheader("Match rate по типам учеников")

    student_stats = (
        df.groupby("student_type")
        .agg(n_turns=("match", "size"), match_rate=("match", "mean"))
        .reset_index()
    )
    student_stats["student_label"] = student_stats["student_type"].map(STUDENT_LABELS).fillna(student_stats["student_type"])
    student_stats["_order"] = student_stats["student_type"].map({s: i for i, s in enumerate(STUDENT_ORDER)}).fillna(99)
    student_stats = student_stats.sort_values("_order", ascending=False)
    student_stats["color"] = student_stats["match_rate"].apply(_match_rate_color)
    student_stats["match_pct"] = (student_stats["match_rate"] * 100).round(1)

    fig_student_bar = go.Figure()
    fig_student_bar.add_trace(
        go.Bar(
            y=student_stats["student_label"],
            x=student_stats["match_pct"],
            orientation="h",
            marker_color=student_stats["color"],
            text=student_stats["match_pct"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>Match rate: %{x:.1f}%<br><extra></extra>",
        )
    )
    fig_student_bar.update_layout(
        xaxis_title="Match rate, %",
        yaxis_title="",
        height=max(300, len(student_stats) * 45),
        margin=dict(l=10, r=40, t=10, b=40),
        xaxis=dict(range=[0, 105]),
        yaxis=dict(categoryorder="array", categoryarray=list(reversed(STUDENT_LABEL_ORDER))),
    )
    fig_student_bar.add_vline(x=80, line_dash="dash", line_color=COLOR_MATCH, annotation_text="80%", annotation_position="top")
    fig_student_bar.add_vline(x=60, line_dash="dash", line_color=COLOR_WARN, annotation_text="60%", annotation_position="top")
    st.plotly_chart(fig_student_bar, use_container_width=True)

    # ── Match rate по типам учеников для каждого интента ──
    st.subheader("Match rate по типам учеников — в разрезе интентов")

    intents_sorted = sorted(df["intent_declared"].dropna().unique())
    cols_per_row = 2
    for row_start in range(0, len(intents_sorted), cols_per_row):
        cols = st.columns(cols_per_row)
        for col_idx, intent in enumerate(intents_sorted[row_start:row_start + cols_per_row]):
            with cols[col_idx]:
                idf = df[df["intent_declared"] == intent]
                ist = (
                    idf.groupby("student_type")
                    .agg(n_turns=("match", "size"), match_rate=("match", "mean"))
                    .reset_index()
                )
                ist["student_label"] = ist["student_type"].map(STUDENT_LABELS).fillna(ist["student_type"])
                ist["color"] = ist["match_rate"].apply(_match_rate_color)
                ist["match_pct"] = (ist["match_rate"] * 100).round(1)

                fig_i = go.Figure()
                fig_i.add_trace(
                    go.Bar(
                        y=ist["student_label"],
                        x=ist["match_pct"],
                        orientation="h",
                        marker_color=ist["color"],
                        text=ist.apply(lambda r: f"{r['match_pct']:.1f}% ({int(r['n_turns'])})", axis=1),
                        textposition="outside",
                        hovertemplate="<b>%{y}</b><br>Match rate: %{x:.1f}%<br><extra></extra>",
                    )
                )
                fig_i.update_layout(
                    title=dict(text=intent, font=dict(size=14)),
                    xaxis_title="",
                    yaxis_title="",
                    height=max(200, len(ist) * 40 + 60),
                    margin=dict(l=10, r=40, t=35, b=20),
                    xaxis=dict(range=[0, 115]),
                    yaxis=dict(categoryorder="array", categoryarray=list(reversed(STUDENT_LABEL_ORDER))),
                )
                fig_i.add_vline(x=80, line_dash="dash", line_color=COLOR_MATCH, opacity=0.4)
                fig_i.add_vline(x=60, line_dash="dash", line_color=COLOR_WARN, opacity=0.4)
                st.plotly_chart(fig_i, use_container_width=True)

    # ── Таблица ──
    st.subheader("Статистика по интентам")

    mismatch_df = df[df["match"] == False]
    if "reason" in mismatch_df.columns and not mismatch_df.empty:
        top_reasons = (
            mismatch_df.groupby("intent_declared")["reason"]
            .apply(lambda s: "; ".join(s.dropna().value_counts().head(3).index.tolist()))
            .reset_index()
            .rename(columns={"reason": "Топ причины несовпадения"})
        )
        table_df = intent_stats.merge(top_reasons, on="intent_declared", how="left")
    else:
        table_df = intent_stats.copy()
        table_df["Топ причины несовпадения"] = ""

    table_df = table_df.rename(
        columns={"intent_declared": "Интент", "n_turns": "Реплик", "match_pct": "Match rate, %"}
    ).sort_values("Match rate, %", ascending=False)

    st.dataframe(
        table_df[["Интент", "Реплик", "Match rate, %", "Топ причины несовпадения"]],
        use_container_width=True,
        hide_index=True,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 2: Модели
# ═══════════════════════════════════════════════════════════════════════════════

with tab_models:
    st.subheader("Сравнение моделей")

    # ── Overall match rate by model ──
    model_stats = (
        df.groupby("model")
        .agg(n_turns=("match", "size"), match_rate=("match", "mean"))
        .reset_index()
        .sort_values("match_rate", ascending=False)
    )
    model_stats["match_pct"] = (model_stats["match_rate"] * 100).round(1)

    fig_model_bar = go.Figure()
    fig_model_bar.add_trace(
        go.Bar(
            x=model_stats["model"],
            y=model_stats["match_pct"],
            marker_color=[MODEL_COLORS.get(m, "#999") for m in model_stats["model"]],
            text=model_stats.apply(lambda r: f"{r['match_pct']:.1f}%<br>({r['n_turns']:,} turns)", axis=1),
            textposition="outside",
        )
    )
    fig_model_bar.update_layout(
        yaxis_title="Match rate, %",
        height=450,
        margin=dict(t=30, b=40),
        yaxis=dict(range=[0, 105]),
    )
    st.plotly_chart(fig_model_bar, use_container_width=True)

    st.markdown("---")

    # ── Match rate by model × student_type ──
    st.subheader("Match rate: модель × тип ученика")

    model_student = (
        df.groupby(["model", "student_type"])
        .agg(n_turns=("match", "size"), match_rate=("match", "mean"))
        .reset_index()
    )
    model_student["match_pct"] = (model_student["match_rate"] * 100).round(1)
    model_student["student_label"] = model_student["student_type"].map(STUDENT_LABELS).fillna(model_student["student_type"])

    fig_grouped = px.bar(
        model_student,
        x="model",
        y="match_pct",
        color="student_label",
        barmode="group",
        text="match_pct",
        color_discrete_sequence=["#2ecc71", "#f39c12", "#e74c3c"],
        category_orders={"student_label": STUDENT_LABEL_ORDER},
        labels={"match_pct": "Match rate, %", "model": "Модель", "student_label": "Тип ученика"},
    )
    fig_grouped.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
    fig_grouped.update_layout(
        height=450,
        margin=dict(t=30, b=40),
        yaxis=dict(range=[0, 105]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
    )
    st.plotly_chart(fig_grouped, use_container_width=True)

    st.markdown("---")

    # ── Heatmap: model × intent ──
    st.subheader("Heatmap: модель × интент")

    heatmap_data = (
        df.groupby(["model", "intent_declared"])["match"]
        .mean()
        .reset_index()
    )
    heatmap_pivot = heatmap_data.pivot(index="model", columns="intent_declared", values="match")
    heatmap_pivot = (heatmap_pivot * 100).round(1)

    # Sort columns by overall match rate
    col_order = df.groupby("intent_declared")["match"].mean().sort_values(ascending=False).index.tolist()
    heatmap_pivot = heatmap_pivot.reindex(columns=[c for c in col_order if c in heatmap_pivot.columns])

    fig_heatmap = go.Figure(
        data=go.Heatmap(
            z=heatmap_pivot.values,
            x=heatmap_pivot.columns.tolist(),
            y=heatmap_pivot.index.tolist(),
            text=heatmap_pivot.values,
            texttemplate="%{text:.1f}%",
            colorscale=[[0, COLOR_MISMATCH], [0.6, COLOR_WARN], [1, COLOR_MATCH]],
            zmin=50,
            zmax=100,
            colorbar=dict(title="Match %"),
        )
    )
    fig_heatmap.update_layout(
        height=max(300, len(heatmap_pivot) * 60 + 100),
        margin=dict(t=20, b=80),
        xaxis=dict(tickangle=-45),
    )
    st.plotly_chart(fig_heatmap, use_container_width=True)

    st.markdown("---")

    # ── Turns per dialog (avg) by model — dialog length comparison ──
    st.subheader("Среднее кол-во реплик на диалог")

    turns_per_dialog = (
        df.groupby(["model", "student_type", "dialog_idx"])
        .size()
        .reset_index(name="n_turns")
        .groupby(["model", "student_type"])["n_turns"]
        .mean()
        .reset_index()
    )
    turns_per_dialog["n_turns"] = turns_per_dialog["n_turns"].round(1)
    turns_per_dialog["student_label"] = turns_per_dialog["student_type"].map(STUDENT_LABELS).fillna(turns_per_dialog["student_type"])

    fig_turns = px.bar(
        turns_per_dialog,
        x="model",
        y="n_turns",
        color="student_label",
        barmode="group",
        text="n_turns",
        color_discrete_sequence=["#2ecc71", "#f39c12", "#e74c3c"],
        category_orders={"student_label": STUDENT_LABEL_ORDER},
        labels={"n_turns": "Реплик / диалог", "model": "Модель", "student_label": "Тип ученика"},
    )
    fig_turns.update_traces(texttemplate="%{text:.1f}", textposition="outside")
    fig_turns.update_layout(
        height=450,
        margin=dict(t=30, b=40),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
    )
    st.plotly_chart(fig_turns, use_container_width=True)

    st.markdown("---")

    # ── Intent distribution by model ──
    st.subheader("Распределение интентов по моделям")

    intent_by_model = (
        df.groupby(["model", "intent_declared"])
        .size()
        .reset_index(name="count")
    )
    # Normalize to percentages within each model
    model_totals = intent_by_model.groupby("model")["count"].transform("sum")
    intent_by_model["pct"] = (intent_by_model["count"] / model_totals * 100).round(1)

    fig_intent_dist = px.bar(
        intent_by_model,
        x="model",
        y="pct",
        color="intent_declared",
        color_discrete_sequence=px.colors.qualitative.Set2,
        labels={"pct": "Доля, %", "model": "Модель", "intent_declared": "Интент"},
    )
    fig_intent_dist.update_layout(
        barmode="stack",
        height=500,
        margin=dict(t=30, b=40),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
    )
    st.plotly_chart(fig_intent_dist, use_container_width=True)

    # ── Сводная таблица ──
    st.subheader("Сводная таблица")

    summary = (
        df.groupby(["model", "student_type"])
        .agg(
            n_turns=("match", "size"),
            match_rate=("match", "mean"),
            avg_confidence=("confidence", "mean"),
            n_mismatch=("match", lambda x: (~x).sum()),
        )
        .reset_index()
    )
    summary["match_rate"] = (summary["match_rate"] * 100).round(1)
    summary["avg_confidence"] = summary["avg_confidence"].round(3)
    summary["student_label"] = summary["student_type"].map(STUDENT_LABELS).fillna(summary["student_type"])
    summary["_order"] = summary["student_type"].map({s: i for i, s in enumerate(STUDENT_ORDER)}).fillna(99)
    summary = summary.sort_values(["model", "_order"])
    summary = summary.rename(columns={
        "model": "Модель",
        "student_label": "Тип ученика",
        "n_turns": "Реплик",
        "match_rate": "Match rate, %",
        "avg_confidence": "Avg confidence",
        "n_mismatch": "Mismatches",
    })
    st.dataframe(
        summary[["Модель", "Тип ученика", "Реплик", "Match rate, %", "Avg confidence", "Mismatches"]],
        use_container_width=True,
        hide_index=True,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 3: Контекст (context appropriateness)
# ═══════════════════════════════════════════════════════════════════════════════

with tab_context:
    if not has_ctx:
        st.info("Данные по контекстуальной уместности отсутствуют. Перезапустите оценку с новым пайплайном.")
    else:
        st.subheader("Уместность интента в контексте диалога")
        st.caption(
            "Проверяет, логично ли использовать данный интент в ответ на реплику репетитора. "
            "Это оценка качества генерации синтетики, а не оценщика."
        )

        ctx_total = len(df)
        ctx_ok = int(df["ctx_appropriate"].sum())
        ctx_bad = ctx_total - ctx_ok

        c1, c2, c3 = st.columns(3)
        c1.metric("Всего реплик", f"{ctx_total:,}")
        c2.metric("Уместных", f"{ctx_ok:,} ({100*ctx_ok/ctx_total:.1f}%)")
        c3.metric("Неуместных", f"{ctx_bad:,} ({100*ctx_bad/ctx_total:.1f}%)")

        st.markdown("---")

        # ── Context appropriateness by intent ──
        st.subheader("Уместность по интентам")
        ctx_by_intent = (
            df.groupby("intent_declared")
            .agg(n=("ctx_appropriate", "size"), appropriate_rate=("ctx_appropriate", "mean"))
            .reset_index()
            .sort_values("appropriate_rate", ascending=True)
        )
        ctx_by_intent["pct"] = (ctx_by_intent["appropriate_rate"] * 100).round(1)
        ctx_by_intent["color"] = ctx_by_intent["appropriate_rate"].apply(_match_rate_color)

        fig_ctx_intent = go.Figure()
        fig_ctx_intent.add_trace(go.Bar(
            y=ctx_by_intent["intent_declared"],
            x=ctx_by_intent["pct"],
            orientation="h",
            marker_color=ctx_by_intent["color"],
            text=ctx_by_intent["pct"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
        ))
        fig_ctx_intent.update_layout(
            xaxis_title="Context appropriate, %",
            height=max(350, len(ctx_by_intent) * 35),
            margin=dict(l=10, r=40, t=10, b=40),
            xaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_ctx_intent, use_container_width=True)

        st.markdown("---")

        # ── Context appropriateness by model ──
        st.subheader("Уместность по моделям")
        ctx_by_model = (
            df.groupby("model")
            .agg(n=("ctx_appropriate", "size"), rate=("ctx_appropriate", "mean"))
            .reset_index()
            .sort_values("rate", ascending=False)
        )
        ctx_by_model["pct"] = (ctx_by_model["rate"] * 100).round(1)

        fig_ctx_model = go.Figure()
        fig_ctx_model.add_trace(go.Bar(
            x=ctx_by_model["model"],
            y=ctx_by_model["pct"],
            marker_color=[MODEL_COLORS.get(m, "#999") for m in ctx_by_model["model"]],
            text=ctx_by_model["pct"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
        ))
        fig_ctx_model.update_layout(
            yaxis_title="Context appropriate, %",
            height=400,
            margin=dict(t=20, b=40),
            yaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_ctx_model, use_container_width=True)

        st.markdown("---")

        # ── Heatmap: model × intent (context) ──
        st.subheader("Heatmap: модель x интент (уместность)")
        ctx_heatmap = (
            df.groupby(["model", "intent_declared"])["ctx_appropriate"]
            .mean().reset_index()
        )
        ctx_pivot = ctx_heatmap.pivot(index="model", columns="intent_declared", values="ctx_appropriate")
        ctx_pivot = (ctx_pivot * 100).round(1)
        col_order = df.groupby("intent_declared")["ctx_appropriate"].mean().sort_values(ascending=False).index.tolist()
        ctx_pivot = ctx_pivot.reindex(columns=[c for c in col_order if c in ctx_pivot.columns])

        fig_ctx_hm = go.Figure(data=go.Heatmap(
            z=ctx_pivot.values,
            x=ctx_pivot.columns.tolist(),
            y=ctx_pivot.index.tolist(),
            text=ctx_pivot.values,
            texttemplate="%{text:.1f}%",
            colorscale=[[0, COLOR_MISMATCH], [0.6, COLOR_WARN], [1, COLOR_MATCH]],
            zmin=50, zmax=100,
            colorbar=dict(title="Appropriate %"),
        ))
        fig_ctx_hm.update_layout(
            height=max(300, len(ctx_pivot) * 60 + 100),
            margin=dict(t=20, b=80),
            xaxis=dict(tickangle=-45),
        )
        st.plotly_chart(fig_ctx_hm, use_container_width=True)

        st.markdown("---")

        # ── Inappropriate turns table ──
        st.subheader("Неуместные реплики")
        ctx_bad_df = df[df["ctx_appropriate"] == False].copy()
        if ctx_bad_df.empty:
            st.success("Все интенты уместны в контексте!")
        else:
            display_cols = [c for c in [
                "model", "student_type", "intent_declared", "student_text",
                "teacher_text_before", "ctx_reason", "match",
            ] if c in ctx_bad_df.columns]
            st.caption(f"{len(ctx_bad_df)} неуместных реплик")
            st.dataframe(
                ctx_bad_df[display_cols].rename(columns={
                    "model": "Модель", "student_type": "Ученик",
                    "intent_declared": "Интент", "student_text": "Текст ученика",
                    "teacher_text_before": "Текст репетитора",
                    "ctx_reason": "Почему неуместен", "match": "Intent match",
                }),
                use_container_width=True, hide_index=True, height=500,
            )

        st.markdown("---")

        # ── Match vs Context quadrant ──
        st.subheader("Квадрант: Intent match vs Context appropriate")
        st.caption("4 группы реплик по двум измерениям оценки")

        q_data = df.copy()
        q_data["quadrant"] = q_data.apply(
            lambda r: (
                "Match + Appropriate" if r["match"] and r["ctx_appropriate"]
                else "Match + Inappropriate" if r["match"] and not r["ctx_appropriate"]
                else "Mismatch + Appropriate" if not r["match"] and r["ctx_appropriate"]
                else "Mismatch + Inappropriate"
            ), axis=1
        )
        q_counts = q_data["quadrant"].value_counts().reset_index()
        q_counts.columns = ["Квадрант", "Кол-во"]
        q_counts["Доля"] = (q_counts["Кол-во"] / q_counts["Кол-во"].sum() * 100).round(1)

        q_colors = {
            "Match + Appropriate": COLOR_MATCH,
            "Match + Inappropriate": COLOR_WARN,
            "Mismatch + Appropriate": "#3498db",
            "Mismatch + Inappropriate": COLOR_MISMATCH,
        }

        fig_q = px.pie(
            q_counts, names="Квадрант", values="Кол-во",
            color="Квадрант",
            color_discrete_map=q_colors,
            hole=0.3,
        )
        fig_q.update_traces(textinfo="label+percent")
        fig_q.update_layout(height=400, margin=dict(t=20, b=20))
        st.plotly_chart(fig_q, use_container_width=True)

        st.dataframe(q_counts, use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# Tab: Ручная проверка
# ═══════════════════════════════════════════════════════════════════════════════

with tab_review:
    st.subheader("Ручная проверка")

    overrides_data = load_overrides()
    n_reviewed = int(df["reviewed"].sum())
    n_overridden = int(df["human_override"].sum())
    n_total = len(df)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Всего реплик", f"{n_total:,}")
    c2.metric("Проверено", f"{n_reviewed:,} ({100*n_reviewed/n_total:.1f}%)" if n_total else "0")
    c3.metric("С правками", f"{n_overridden:,}")
    c4.metric("Без проверки", f"{n_total - n_reviewed:,}")

    st.progress(n_reviewed / n_total if n_total else 0, text=f"Прогресс проверки: {n_reviewed}/{n_total}")

    st.markdown("---")

    if n_reviewed > 0:
        # Accuracy of model vs human
        reviewed_df = df[df["reviewed"] == True].copy()

        st.subheader("Точность модели (на проверенных)")

        if n_overridden > 0:
            overridden_df = df[df["human_override"] == True].copy()

            # How many overrides changed the match verdict?
            match_changes = 0
            ctx_changes = 0
            for _, row in overridden_df.iterrows():
                key = _turn_key(row)
                if key in overrides_data:
                    ov = overrides_data[key]
                    # We can't easily compare original vs override since we already applied it
                    # But we know human_override=True means something was changed
                    match_changes += 1

            st.markdown(f"**{n_overridden}** реплик исправлены вручную из **{n_reviewed}** проверенных "
                        f"({100*n_overridden/n_reviewed:.1f}% ошибок модели)")

            # Breakdown of corrections by intent
            st.subheader("Правки по интентам")
            corrections_by_intent = (
                overridden_df.groupby("intent_declared")
                .size()
                .reset_index(name="corrections")
            )
            reviewed_by_intent = (
                reviewed_df.groupby("intent_declared")
                .size()
                .reset_index(name="reviewed")
            )
            intent_review = reviewed_by_intent.merge(corrections_by_intent, on="intent_declared", how="left").fillna(0)
            intent_review["corrections"] = intent_review["corrections"].astype(int)
            intent_review["error_rate"] = (intent_review["corrections"] / intent_review["reviewed"] * 100).round(1)
            intent_review = intent_review.sort_values("error_rate", ascending=False)

            fig_review = go.Figure()
            fig_review.add_trace(go.Bar(
                y=intent_review["intent_declared"],
                x=intent_review["error_rate"],
                orientation="h",
                marker_color=intent_review["error_rate"].apply(
                    lambda v: COLOR_MISMATCH if v > 20 else COLOR_WARN if v > 10 else COLOR_MATCH
                ),
                text=intent_review.apply(
                    lambda r: f"{r['error_rate']:.1f}% ({r['corrections']}/{r['reviewed']})", axis=1
                ),
                textposition="outside",
            ))
            fig_review.update_layout(
                xaxis_title="Error rate модели, %",
                height=max(300, len(intent_review) * 35),
                margin=dict(l=10, r=80, t=10, b=40),
            )
            st.plotly_chart(fig_review, use_container_width=True)

            # Corrections by model
            st.subheader("Правки по моделям")
            corrections_by_model = (
                overridden_df.groupby("model")
                .size()
                .reset_index(name="corrections")
            )
            reviewed_by_model = (
                reviewed_df.groupby("model")
                .size()
                .reset_index(name="reviewed")
            )
            model_review = reviewed_by_model.merge(corrections_by_model, on="model", how="left").fillna(0)
            model_review["corrections"] = model_review["corrections"].astype(int)
            model_review["error_rate"] = (model_review["corrections"] / model_review["reviewed"] * 100).round(1)

            fig_mr = go.Figure()
            fig_mr.add_trace(go.Bar(
                x=model_review["model"],
                y=model_review["error_rate"],
                marker_color=[MODEL_COLORS.get(m, "#999") for m in model_review["model"]],
                text=model_review.apply(
                    lambda r: f"{r['error_rate']:.1f}% ({r['corrections']}/{r['reviewed']})", axis=1
                ),
                textposition="outside",
            ))
            fig_mr.update_layout(
                yaxis_title="Error rate модели, %",
                height=400, margin=dict(t=20, b=40),
            )
            st.plotly_chart(fig_mr, use_container_width=True)
        else:
            st.success("Все проверенные реплики совпали с оценкой модели — правок нет.")

        st.markdown("---")

        # Table of all corrections
        st.subheader("Все ручные правки")
        if n_overridden > 0:
            override_rows = []
            for key, ov in overrides_data.items():
                if not ov.get("human_reason") and ov.get("match") is None:
                    continue
                parts = key.split("|")
                if len(parts) == 4:
                    override_rows.append({
                        "Модель": parts[0],
                        "Ученик": STUDENT_LABELS.get(parts[1], parts[1]),
                        "Диалог": parts[2],
                        "Ход": parts[3],
                        "Match": ov.get("match", "—"),
                        "Ctx": ov.get("ctx_appropriate", "—"),
                        "Интент": ov.get("actual_intent", ""),
                        "Комментарий": ov.get("human_reason", ""),
                    })
            if override_rows:
                st.dataframe(pd.DataFrame(override_rows), use_container_width=True, hide_index=True)
        else:
            st.info("Нет ручных правок.")

    else:
        st.info("Пока ни одна реплика не проверена. Откройте таб «Диалоги» и отмечайте чекбокс «Проверено».")

    st.markdown("---")

    # Review coverage by model/student
    st.subheader("Покрытие проверки")
    review_coverage = (
        df.groupby(["model", "student_type"])
        .agg(total=("reviewed", "size"), reviewed=("reviewed", "sum"))
        .reset_index()
    )
    review_coverage["reviewed"] = review_coverage["reviewed"].astype(int)
    review_coverage["pct"] = (review_coverage["reviewed"] / review_coverage["total"] * 100).round(1)
    review_coverage["student_label"] = review_coverage["student_type"].map(STUDENT_LABELS).fillna(review_coverage["student_type"])
    review_coverage["_order"] = review_coverage["student_type"].map({s: i for i, s in enumerate(STUDENT_ORDER)}).fillna(99)
    review_coverage = review_coverage.sort_values(["model", "_order"])
    st.dataframe(
        review_coverage[["model", "student_label", "total", "reviewed", "pct"]].rename(columns={
            "model": "Модель", "student_label": "Ученик", "total": "Всего",
            "reviewed": "Проверено", "pct": "Покрытие, %",
        }),
        use_container_width=True, hide_index=True,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 5: Распределение
# ═══════════════════════════════════════════════════════════════════════════════

with tab_distrib:
    # ── Общее: фактическое распределение vs дефолтные веса ──
    st.subheader("Фактическое распределение интентов vs дефолтные веса (все ученики)")

    actual_counts_all = df["intent_declared"].value_counts()
    actual_pct_all = (actual_counts_all / actual_counts_all.sum() * 100).reset_index()
    actual_pct_all.columns = ["intent", "actual_pct"]

    weights_df = pd.DataFrame(list(DEFAULT_WEIGHTS.items()), columns=["intent", "default_weight"])
    weights_df["default_pct"] = weights_df["default_weight"] / weights_df["default_weight"].sum() * 100

    compare_all = weights_df.merge(actual_pct_all, on="intent", how="outer").fillna(0)
    compare_all["delta"] = (compare_all["actual_pct"] - compare_all["default_pct"]).round(2)
    compare_all = compare_all.sort_values("default_pct", ascending=False)

    fig_dist_all = go.Figure()
    fig_dist_all.add_trace(
        go.Bar(name="Дефолтные веса, %", x=compare_all["intent"], y=compare_all["default_pct"], marker_color="#3498db")
    )
    fig_dist_all.add_trace(
        go.Bar(name="Фактическое, %", x=compare_all["intent"], y=compare_all["actual_pct"], marker_color="#e67e22")
    )
    fig_dist_all.update_layout(
        barmode="group",
        xaxis_title="Интент",
        yaxis_title="Доля, %",
        height=500,
        margin=dict(t=30, b=80),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
    )
    st.plotly_chart(fig_dist_all, use_container_width=True, key="distrib_all")

    st.markdown("---")

    # ── Распределение интентов по каждому типу ученика ──
    st.subheader("Распределение интентов по типам учеников")

    student_types_sorted = [s for s in STUDENT_ORDER if s in df["student_type"].values]
    for stype in student_types_sorted:
        slabel = STUDENT_LABELS.get(stype, stype)
        sdf = df[df["student_type"] == stype]

        s_counts = sdf["intent_declared"].value_counts()
        s_pct = (s_counts / s_counts.sum() * 100).reset_index()
        s_pct.columns = ["intent", "actual_pct"]

        s_compare = weights_df[["intent", "default_pct"]].merge(s_pct, on="intent", how="outer").fillna(0)
        s_compare["delta"] = (s_compare["actual_pct"] - s_compare["default_pct"]).round(2)
        s_compare = s_compare.sort_values("default_pct", ascending=False)

        fig_s = go.Figure()
        fig_s.add_trace(
            go.Bar(name="Дефолтные веса, %", x=s_compare["intent"], y=s_compare["default_pct"], marker_color="#3498db")
        )
        fig_s.add_trace(
            go.Bar(name="Фактическое, %", x=s_compare["intent"], y=s_compare["actual_pct"], marker_color="#e67e22")
        )
        fig_s.update_layout(
            title=dict(text=f"{slabel} ({len(sdf)} реплик)", font=dict(size=16)),
            barmode="group",
            xaxis_title="Интент",
            yaxis_title="Доля, %",
            height=400,
            margin=dict(t=40, b=80),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        )
        st.plotly_chart(fig_s, use_container_width=True, key=f"distrib_{stype}")

    st.markdown("---")

    # ── Сводная таблица расхождений по типам учеников ──
    st.subheader("Расхождения (факт − дефолт, п.п.)")

    pivot_rows = []
    for stype in student_types_sorted:
        slabel = STUDENT_LABELS.get(stype, stype)
        sdf = df[df["student_type"] == stype]
        s_counts = sdf["intent_declared"].value_counts()
        s_pct = (s_counts / s_counts.sum() * 100)
        for intent in weights_df["intent"]:
            default_p = weights_df.loc[weights_df["intent"] == intent, "default_pct"].values[0]
            actual_p = s_pct.get(intent, 0.0)
            pivot_rows.append({
                "Тип ученика": slabel,
                "Интент": intent,
                "Дефолт, %": round(default_p, 1),
                "Факт, %": round(actual_p, 1),
                "Дельта, п.п.": round(actual_p - default_p, 1),
            })

    pivot_df = pd.DataFrame(pivot_rows)
    delta_pivot = pivot_df.pivot(index="Интент", columns="Тип ученика", values="Дельта, п.п.").reset_index()
    delta_pivot = delta_pivot.reindex(
        columns=["Интент"] + [STUDENT_LABELS.get(s, s) for s in student_types_sorted]
    )

    st.dataframe(
        delta_pivot.style.map(
            lambda v: f"color: {COLOR_MISMATCH}" if isinstance(v, (int, float)) and abs(v) > 5 else "",
        ),
        use_container_width=True,
        hide_index=True,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 4: Проблемные реплики
# ═══════════════════════════════════════════════════════════════════════════════

with tab_problems:
    st.subheader("Реплики с match = False")

    problem_df = df[df["match"] == False].copy()

    if problem_df.empty:
        st.success("Нет проблемных реплик — все интенты совпали!")
    else:
        col_f1, col_f2 = st.columns(2)

        with col_f1:
            all_intents = sorted(problem_df["intent_declared"].dropna().unique().tolist())
            selected_intents = st.multiselect(
                "Фильтр по интенту",
                options=all_intents,
                default=all_intents,
                key="problem_intent_filter",
            )

        with col_f2:
            if "confidence" in problem_df.columns and problem_df["confidence"].notna().any():
                conf_min = float(problem_df["confidence"].min())
                conf_max = float(problem_df["confidence"].max())
                if conf_min < conf_max:
                    conf_thresh = st.slider(
                        "Порог уверенности (показать <=)",
                        min_value=conf_min,
                        max_value=conf_max,
                        value=conf_max,
                        step=0.05,
                        key="conf_slider",
                    )
                else:
                    st.info(f"Все confidence = {conf_max:.2f}")
                    conf_thresh = conf_max
            else:
                conf_thresh = None

        filtered = problem_df[problem_df["intent_declared"].isin(selected_intents)]
        if conf_thresh is not None and "confidence" in filtered.columns:
            filtered = filtered[filtered["confidence"] <= conf_thresh]

        display_cols = [
            c
            for c in ["model", "student_type", "intent_declared", "student_text",
                       "teacher_text_before", "reason", "actual_intent", "confidence"]
            if c in filtered.columns
        ]

        col_rename = {
            "model": "Модель",
            "student_type": "Тип ученика",
            "intent_declared": "Объявленный интент",
            "student_text": "Текст ученика",
            "teacher_text_before": "Текст репетитора до",
            "reason": "Причина",
            "actual_intent": "Фактический интент",
            "confidence": "Уверенность",
        }

        st.caption(f"Показано {len(filtered)} из {len(problem_df)} проблемных реплик")
        st.dataframe(
            filtered[display_cols].rename(columns=col_rename),
            use_container_width=True,
            hide_index=True,
            height=600,
        )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 5: Answer детально
# ═══════════════════════════════════════════════════════════════════════════════

with tab_answer:
    st.subheader("Глубокий анализ интента answer")

    answer_df = df[df["intent_declared"] == "answer"].copy()

    if answer_df.empty:
        st.info("Нет данных по интенту answer.")
    else:
        n_answer = len(answer_df)
        n_match = answer_df["match"].sum()
        n_mismatch = n_answer - n_match
        rate = n_match / n_answer if n_answer > 0 else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Всего реплик answer", f"{n_answer:,}")
        c2.metric("Совпало", f"{n_match:,}")
        c3.metric("Не совпало", f"{n_mismatch:,}")
        c4.metric("Match rate", f"{rate:.1%}")

        st.markdown("---")

        # Answer match rate by model
        st.subheader("Answer match rate по моделям")
        answer_by_model = (
            answer_df.groupby("model")
            .agg(n=("match", "size"), match_rate=("match", "mean"))
            .reset_index()
        )
        answer_by_model["match_pct"] = (answer_by_model["match_rate"] * 100).round(1)

        fig_answer_model = go.Figure()
        fig_answer_model.add_trace(
            go.Bar(
                x=answer_by_model["model"],
                y=answer_by_model["match_pct"],
                marker_color=[MODEL_COLORS.get(m, "#999") for m in answer_by_model["model"]],
                text=answer_by_model["match_pct"].apply(lambda v: f"{v:.1f}%"),
                textposition="outside",
            )
        )
        fig_answer_model.update_layout(
            yaxis_title="Match rate, %",
            height=400,
            margin=dict(t=20, b=40),
            yaxis=dict(range=[0, 105]),
        )
        st.plotly_chart(fig_answer_model, use_container_width=True)

        st.markdown("---")

        # What are mismatched answers actually?
        answer_mismatch = answer_df[answer_df["match"] == False]

        if "actual_intent" in answer_mismatch.columns and answer_mismatch["actual_intent"].notna().any():
            st.subheader("Чем на самом деле являются несовпавшие answer?")

            actual_dist = answer_mismatch["actual_intent"].value_counts().reset_index()
            actual_dist.columns = ["actual_intent", "count"]

            fig_pie = px.pie(
                actual_dist,
                names="actual_intent",
                values="count",
                color_discrete_sequence=px.colors.qualitative.Set2,
                hole=0.3,
            )
            fig_pie.update_layout(height=450, margin=dict(t=20, b=20))
            fig_pie.update_traces(
                textinfo="label+percent",
                hovertemplate="<b>%{label}</b><br>Кол-во: %{value}<br>Доля: %{percent}<extra></extra>",
            )
            st.plotly_chart(fig_pie, use_container_width=True)

            actual_dist["pct"] = (actual_dist["count"] / actual_dist["count"].sum() * 100).round(1)
            actual_dist.columns = ["Фактический интент", "Кол-во", "Доля, %"]
            st.dataframe(actual_dist, use_container_width=True, hide_index=True)

        # Confidence distribution for answer
        if "confidence" in answer_df.columns and answer_df["confidence"].notna().any():
            st.subheader("Распределение уверенности (answer)")
            fig_hist = px.histogram(
                answer_df,
                x="confidence",
                color=answer_df["match"].map({True: "Match", False: "Mismatch"}),
                color_discrete_map={"Match": COLOR_MATCH, "Mismatch": COLOR_MISMATCH},
                nbins=20,
                barmode="overlay",
                opacity=0.7,
                labels={"color": "Результат", "confidence": "Уверенность"},
            )
            fig_hist.update_layout(height=350, margin=dict(t=20, b=40))
            st.plotly_chart(fig_hist, use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 6: Диалоги
# ═══════════════════════════════════════════════════════════════════════════════

with tab_dialogs:
    st.subheader("Просмотр диалогов с результатами оценки")

    source_df = load_source_dialogs(_mtime(RESULTS_PATH))

    if source_df is None:
        st.warning("Файлы с диалогами не найдены в `data/intent_eval/`")
    else:
        # Filter source_df by sidebar selections
        source_filtered = source_df[
            source_df["model"].isin(selected_models) & source_df["student_type"].isin(selected_students)
        ].copy()

        # Get evaluated dialog indices (matching model+student_type+dialog_idx)
        eval_keys = set(
            df[["model", "student_type", "dialog_idx"]].apply(
                lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1
            )
        )
        source_evaluated = source_filtered[
            source_filtered.apply(
                lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in eval_keys, axis=1
            )
        ].copy()

        if source_evaluated.empty:
            st.info("Нет диалогов, совпадающих с результатами оценки.")
        else:
            # Filters
            col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
            with col_f1:
                filter_result = st.selectbox(
                    "Результат",
                    ["Все", "Есть mismatch", "Все match"],
                    key="dialog_filter_result",
                )
            with col_f2:
                available_intents = sorted(df["intent_declared"].dropna().unique().tolist())
                filter_intent = st.selectbox(
                    "Интент",
                    ["Все"] + available_intents,
                    key="dialog_filter_intent",
                )
            with col_f3:
                filter_ctx = st.selectbox(
                    "Контекст",
                    ["Все", "Есть неуместные"],
                    key="dialog_filter_ctx",
                )
            with col_f4:
                filter_defect = st.selectbox(
                    "Брак",
                    ["Все", "Есть брак", "Без брака"],
                    key="dialog_filter_defect",
                )
            with col_f5:
                filter_reviewed = st.selectbox(
                    "Проверка",
                    ["Все", "Не проверенные", "Проверенные"],
                    key="dialog_filter_reviewed",
                )

            # Apply filters to get relevant dialog keys
            filtered_results = df.copy()
            if filter_intent != "Все":
                filtered_results = filtered_results[filtered_results["intent_declared"] == filter_intent]
            if filter_result == "Есть mismatch":
                mismatch_keys = set(
                    filtered_results[filtered_results["match"] == False][["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in mismatch_keys, axis=1
                    )
                ]
            elif filter_result == "Все match":
                all_keys = set(
                    filtered_results[["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                mismatch_keys = set(
                    filtered_results[filtered_results["match"] == False][["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                ok_keys = all_keys - mismatch_keys
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in ok_keys, axis=1
                    )
                ]

            # Context appropriateness filter
            if filter_ctx == "Есть неуместные" and has_ctx:
                ctx_bad_keys = set(
                    filtered_results[filtered_results["ctx_appropriate"] == False][["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in ctx_bad_keys, axis=1
                    )
                ]

            # Defect filter
            if filter_defect != "Все" and has_defect:
                if filter_defect == "Есть брак":
                    defect_keys = set(
                        filtered_results[filtered_results["is_defect"] == True][["model", "student_type", "dialog_idx"]]
                        .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                    )
                else:  # Без брака
                    all_d_keys = set(
                        filtered_results[["model", "student_type", "dialog_idx"]]
                        .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                    )
                    defect_d_keys = set(
                        filtered_results[filtered_results["is_defect"] == True][["model", "student_type", "dialog_idx"]]
                        .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                    )
                    defect_keys = all_d_keys - defect_d_keys
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in defect_keys, axis=1
                    )
                ]

            # Reviewed filter
            if filter_reviewed == "Не проверенные" and "reviewed" in filtered_results.columns:
                unreviewed_keys = set(
                    filtered_results[filtered_results["reviewed"] == False][["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in unreviewed_keys, axis=1
                    )
                ]
            elif filter_reviewed == "Проверенные" and "reviewed" in filtered_results.columns:
                reviewed_keys = set(
                    filtered_results[filtered_results["reviewed"] == True][["model", "student_type", "dialog_idx"]]
                    .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                )
                filtered_results = filtered_results[
                    filtered_results.apply(
                        lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}" in reviewed_keys, axis=1
                    )
                ]

            # Build visible dialog list
            if filtered_results.empty:
                visible_keys = []
            else:
                _sorder = {s: i for i, s in enumerate(STUDENT_ORDER)}
                visible_keys = sorted(
                    set(
                        filtered_results[["model", "student_type", "dialog_idx"]]
                        .apply(lambda r: f"{r['model']}|{r['student_type']}|{r['dialog_idx']}", axis=1)
                    ),
                    key=lambda k: (k.split("|")[0], _sorder.get(k.split("|")[1], 99), int(k.split("|")[2])),
                )

            if not visible_keys:
                st.info("Нет диалогов по заданным фильтрам.")
            else:
                # Build labels
                dialog_options = []
                dialog_labels = {}
                for key in visible_keys:
                    parts = key.split("|")
                    m, s, idx = parts[0], parts[1], int(parts[2])
                    d_results = df[(df["model"] == m) & (df["student_type"] == s) & (df["dialog_idx"] == idx)]
                    n_match = int(d_results["match"].sum())
                    n_total = len(d_results)
                    status = "✅" if n_match == n_total else "⚠️"
                    s_label = STUDENT_LABELS.get(s, s)
                    label = f"{status} {m} / {s_label} / Диалог {idx} — {n_match}/{n_total} match"
                    dialog_options.append(key)
                    dialog_labels[key] = label

                selected_key = st.selectbox(
                    f"Выберите диалог ({len(visible_keys)} шт.)",
                    options=dialog_options,
                    format_func=lambda x: dialog_labels.get(x, x),
                    key="dialog_selector",
                )

                # Parse selected key
                sel_parts = selected_key.split("|")
                sel_model, sel_student, sel_idx = sel_parts[0], sel_parts[1], int(sel_parts[2])

                # Get dialog text from source
                dialog_row = source_evaluated[
                    (source_evaluated["model"] == sel_model) &
                    (source_evaluated["student_type"] == sel_student) &
                    (source_evaluated["dialog_idx"] == sel_idx)
                ]
                if dialog_row.empty:
                    st.error("Диалог не найден в исходных файлах.")
                else:
                    dialog_row = dialog_row.iloc[0]
                    dialog_results = df[
                        (df["model"] == sel_model) &
                        (df["student_type"] == sel_student) &
                        (df["dialog_idx"] == sel_idx)
                    ].copy()

                    col_i1, col_i2, col_i3, col_i4 = st.columns(4)
                    col_i1.metric("Модель", sel_model)
                    col_i2.metric("Ученик", STUDENT_LABELS.get(sel_student, sel_student))
                    col_i3.metric("Класс", dialog_row["grade_group"])
                    n_m = int(dialog_results["match"].sum())
                    n_t = len(dialog_results)
                    col_i4.metric("Match rate", f"{n_m}/{n_t}" if n_t else "—")

                    st.markdown("---")

                    # Build eval lookup
                    eval_lookup: dict[int, dict] = {}
                    for _, r in dialog_results.iterrows():
                        eval_lookup[int(r["turn_idx"])] = r.to_dict()

                    # Render dialog
                    turns = _parse_dialog_to_turns(dialog_row["dialog"])
                    user_turn_counter = 0

                    for turn in turns:
                        if turn["role"] == "user":
                            user_turn_counter += 1

                        if turn["role"] == "assistant":
                            st.markdown(
                                f'<div style="background-color:#1e3a5f; padding:10px 14px; '
                                f'border-radius:8px; margin:4px 0; border-left:4px solid #3498db;">'
                                f'<b>🧑‍🏫 Репетитор</b><br>{turn["text"]}</div>',
                                unsafe_allow_html=True,
                            )
                        elif turn["role"] == "user":
                            intent_tag = f" [{turn['intent']}]" if turn.get("intent") else ""
                            eval_info = eval_lookup.get(user_turn_counter)

                            if eval_info is not None:
                                is_match = eval_info["match"]
                                bg = "#1a3d1a" if is_match else "#4d1a1a"
                                border = COLOR_MATCH if is_match else COLOR_MISMATCH
                                icon = "✅" if is_match else "❌"
                                ctx_ok = eval_info.get("ctx_appropriate", True)
                                is_defect_turn = eval_info.get("is_defect", False)
                                defect_badge = ' <span style="font-size:0.8em; background:#E74C3C; padding:1px 5px; border-radius:3px;">БРАК</span>' if is_defect_turn else ""
                                ctx_badge = "" if ctx_ok else ' <span style="font-size:0.8em; background:#8B4513; padding:1px 5px; border-radius:3px;">ctx: неуместен</span>'
                                human_badge = ' <span style="font-size:0.8em; background:#2E86C1; padding:1px 5px; border-radius:3px;">human</span>' if eval_info.get("human_override") else ""
                                reviewed_badge = ' <span style="font-size:0.8em; background:#27AE60; padding:1px 5px; border-radius:3px;">reviewed</span>' if eval_info.get("reviewed") else ""
                                badge = f' <span style="font-size:0.8em; opacity:0.8;">{icon} conf={eval_info["confidence"]:.2f}</span>{defect_badge}{ctx_badge}{human_badge}{reviewed_badge}'
                                details = ""
                                if eval_info.get("reason"):
                                    actual = eval_info.get("actual_intent", "")
                                    actual_str = f" → <b>{actual}</b>" if actual else ""
                                    details = (
                                        f'<div style="font-size:0.85em; margin-top:6px; padding:6px 8px; '
                                        f'background:rgba(0,0,0,0.2); border-radius:4px;">'
                                        f'💬 {eval_info["reason"]}{actual_str}</div>'
                                    )
                            else:
                                bg = "#2a2a2a"
                                border = "#666"
                                badge = ""
                                details = ""

                            st.markdown(
                                f'<div style="background-color:{bg}; padding:10px 14px; '
                                f'border-radius:8px; margin:4px 0; border-left:4px solid {border};">'
                                f'<b>👤 Ученик{intent_tag}</b>{badge}<br>{turn["text"]}'
                                f'{details}</div>',
                                unsafe_allow_html=True,
                            )

                            # Quick review checkbox (outside expander)
                            if eval_info is not None:
                                _ok = _turn_key({
                                    "model": sel_model, "student_type": sel_student,
                                    "dialog_idx": sel_idx, "turn_idx": user_turn_counter,
                                })
                                _ovs = load_overrides()
                                _is_rev = _ovs.get(_ok, {}).get("reviewed", False)
                                _chk = f"chk_inline_{sel_model}_{sel_student}_{sel_idx}_{user_turn_counter}"
                                _new_rev = st.checkbox("Проверено", value=_is_rev, key=_chk)
                                if _new_rev != _is_rev:
                                    if _ok not in _ovs:
                                        _ovs[_ok] = {}
                                    _ovs[_ok]["reviewed"] = _new_rev
                                    save_overrides(_ovs)
                                    st.cache_data.clear()
                                    st.rerun()

                            # Decision log expander
                            if eval_info is not None and eval_info.get("eval_1"):
                                with st.expander(f"🔍 Лог решения (ход {user_turn_counter})", expanded=False):
                                    def _parse_eval(raw):
                                        if not raw or pd.isna(raw):
                                            return {}
                                        try:
                                            return json.loads(raw) if isinstance(raw, str) else {}
                                        except (json.JSONDecodeError, TypeError):
                                            return {}

                                    for i, (label, temp) in enumerate([
                                        ("Эксперт 1", "t=0.2"), ("Эксперт 2", "t=0.5"), ("Эксперт 3", "t=0.8")
                                    ], 1):
                                        ev = _parse_eval(eval_info.get(f"eval_{i}", ""))
                                        if ev:
                                            ev_icon = "✅" if ev.get("match") else "❌"
                                            ev_actual = f" → {ev.get('actual_intent')}" if ev.get("actual_intent") else ""
                                            st.markdown(
                                                f"**{label}** ({temp}): {ev_icon} match={ev.get('match')} "
                                                f"conf={ev.get('confidence', '?')}{ev_actual}\n\n"
                                                f"> {ev.get('reason', '—')}"
                                            )
                                        else:
                                            st.markdown(f"**{label}** ({temp}): ⚠️ нет данных")

                                    judge = eval_info.get("judge_reasoning", "")
                                    if judge and not pd.isna(judge):
                                        st.markdown(f"---\n**⚖️ Судья (intent):** {judge}")

                                    # Context appropriateness section
                                    if eval_info.get("ctx_eval_1") and not pd.isna(eval_info.get("ctx_eval_1", "")):
                                        ctx_ok = eval_info.get("ctx_appropriate", True)
                                        ctx_icon = "✅" if ctx_ok else "⚠️"
                                        st.markdown(f"---\n**{ctx_icon} Контекстуальная уместность:** {'уместен' if ctx_ok else 'НЕУМЕСТЕН'}")
                                        if eval_info.get("ctx_reason") and not pd.isna(eval_info.get("ctx_reason", "")):
                                            st.markdown(f"> {eval_info['ctx_reason']}")

                                        for i, (label, temp) in enumerate([
                                            ("Эксперт 1", "t=0.2"), ("Эксперт 2", "t=0.5"), ("Эксперт 3", "t=0.8")
                                        ], 1):
                                            cev = _parse_eval(eval_info.get(f"ctx_eval_{i}", ""))
                                            if cev:
                                                c_icon = "✅" if cev.get("appropriate") else "⚠️"
                                                st.markdown(
                                                    f"**{label}** ({temp}): {c_icon} appropriate={cev.get('appropriate')}\n\n"
                                                    f"> {cev.get('reason', '—')}"
                                                )

                                        ctx_j = eval_info.get("ctx_judge", "")
                                        if ctx_j and not pd.isna(ctx_j) and "skipped" not in str(ctx_j).lower():
                                            st.markdown(f"**⚖️ Судья (контекст):** {ctx_j}")

                                    # Human review section
                                    st.markdown("---")
                                    override_key = _turn_key({
                                        "model": sel_model, "student_type": sel_student,
                                        "dialog_idx": sel_idx, "turn_idx": user_turn_counter,
                                    })
                                    current_overrides = load_overrides()
                                    ov_existing = current_overrides.get(override_key, {})
                                    has_override = bool(ov_existing.get("match") is not None or ov_existing.get("human_reason"))
                                    is_reviewed = ov_existing.get("reviewed", False)

                                    # Review status (read-only display, checkbox is above)
                                    if is_reviewed:
                                        st.markdown("✅ **Проверено**")

                                    if has_override:
                                        ov = current_overrides[override_key]
                                        st.markdown(
                                            f"**✏️ Ручная правка:** match={ov.get('match', '—')}, "
                                            f"ctx={ov.get('ctx_appropriate', '—')}, "
                                            f"intent={ov.get('actual_intent', '—')}\n\n"
                                            f"> {ov.get('human_reason', '')}"
                                        )

                                    form_key = f"override_{sel_model}_{sel_student}_{sel_idx}_{user_turn_counter}"
                                    with st.form(key=form_key):
                                        st.markdown("**✏️ Исправить вердикт**" if not has_override else "**✏️ Обновить правку**")

                                        ov_defect = st.selectbox(
                                            "Брак",
                                            [False, True],
                                            index=1 if ov_existing.get("is_defect", eval_info.get("is_defect", False)) else 0,
                                            format_func=lambda x: "Да — брак" if x else "Нет",
                                            key=f"{form_key}_defect",
                                        )
                                        ov_match = st.selectbox(
                                            "Intent match",
                                            [True, False],
                                            index=0 if ov_existing.get("match", eval_info.get("match", True)) else 1,
                                            key=f"{form_key}_match",
                                        )
                                        ov_ctx = st.selectbox(
                                            "Context appropriate",
                                            [True, False],
                                            index=0 if ov_existing.get("ctx_appropriate", eval_info.get("ctx_appropriate", True)) else 1,
                                            key=f"{form_key}_ctx",
                                        )
                                        all_intents = ["", "answer", "get-explanation", "get-solution",
                                                       "agree-with-tutor", "chat", "thank-tutor",
                                                       "set-problem", "criticize-tutor"]
                                        default_actual = ov_existing.get("actual_intent", str(eval_info.get("actual_intent", "")))
                                        actual_idx = all_intents.index(default_actual) if default_actual in all_intents else 0
                                        ov_actual = st.selectbox(
                                            "Фактический интент (если отличается)",
                                            all_intents,
                                            index=actual_idx,
                                            key=f"{form_key}_actual",
                                        )
                                        ov_reason = st.text_input(
                                            "Комментарий",
                                            value=ov_existing.get("human_reason", ""),
                                            key=f"{form_key}_reason",
                                        )

                                        col_save, col_delete = st.columns(2)
                                        submitted = col_save.form_submit_button("Сохранить")
                                        deleted = col_delete.form_submit_button("Удалить правку") if has_override else None

                                    if submitted:
                                        current_overrides[override_key] = {
                                            "is_defect": ov_defect,
                                            "match": ov_match,
                                            "ctx_appropriate": ov_ctx,
                                            "actual_intent": ov_actual,
                                            "human_reason": ov_reason,
                                            "reviewed": True,
                                        }
                                        save_overrides(current_overrides)
                                        st.cache_data.clear()
                                        st.rerun()

                                    if deleted:
                                        current_overrides.pop(override_key, None)
                                        save_overrides(current_overrides)
                                        st.cache_data.clear()
                                        st.rerun()

                    # Summary table
                    if not dialog_results.empty:
                        with st.expander("📋 Таблица результатов по репликам"):
                            show_cols = [c for c in ["turn_idx", "intent_declared", "match", "confidence", "reason", "actual_intent"]
                                         if c in dialog_results.columns]
                            st.dataframe(
                                dialog_results[show_cols].rename(columns={
                                    "turn_idx": "Ход", "intent_declared": "Интент",
                                    "match": "Match", "confidence": "Уверенность",
                                    "reason": "Причина", "actual_intent": "Факт. интент",
                                }),
                                use_container_width=True, hide_index=True,
                            )

# ═══════════════════════════════════════════════════════════════════════════════
# Tab 7: Аудит
# ═══════════════════════════════════════════════════════════════════════════════

with tab_audit:
    st.subheader("Аудит-таблица по интентам")
    st.caption("Заполните данные аудита и сохраните в CSV.")

    audit_df = pd.DataFrame(AUDIT_INTENTS)
    audit_df = audit_df.rename(
        columns={
            "id": "Intent ID",
            "name": "Название",
            "default_weight": "Дефолтный вес",
            "prompt_file": "Промпт-файл",
        }
    )

    audit_df["Кто работал"] = ""
    audit_df["Что замеряли"] = ""
    audit_df["Размер пула"] = 0
    audit_df["Match rate, %"] = 0.0
    audit_df["Промпт переработан"] = False
    audit_df["Дельта, п.п."] = 0.0
    audit_df["Комментарии"] = ""

    edited = st.data_editor(
        audit_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "Intent ID": st.column_config.TextColumn(disabled=True),
            "Название": st.column_config.TextColumn(disabled=True),
            "Дефолтный вес": st.column_config.NumberColumn(disabled=True),
            "Промпт-файл": st.column_config.TextColumn(disabled=True),
            "Кто работал": st.column_config.TextColumn(width="medium"),
            "Что замеряли": st.column_config.TextColumn(width="medium"),
            "Размер пула": st.column_config.NumberColumn(min_value=0, step=1),
            "Match rate, %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.1, format="%.1f"),
            "Промпт переработан": st.column_config.CheckboxColumn(),
            "Дельта, п.п.": st.column_config.NumberColumn(step=0.1, format="%.1f"),
            "Комментарии": st.column_config.TextColumn(width="large"),
        },
        key="audit_editor",
    )

    export_path = DATA_DIR / "audit_intents.csv"

    if st.button("💾 Сохранить аудит в CSV", type="primary"):
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        edited.to_csv(export_path, index=False, encoding="utf-8-sig")
        st.success(f"Сохранено: `{export_path}`")

# ═══════════════════════════════════════════════════════════════════════════════
# Tab: Справка
# ═══════════════════════════════════════════════════════════════════════════════

with tab_reference:
    st.subheader("Справка: пайплайн оценки интентов")

    st.markdown("""
Пайплайн оценивает каждую реплику ученика по **3 критериям** с помощью ансамбля из 2 экспертов + судьи:

1. **Брак (is_defect)** — английский текст, протекание промпта, отказ модели, слишком длинный текст
2. **Соответствие интенту (match)** — соответствует ли текст реплики объявленному интенту
3. **Контекстная уместность (ctx_appropriate)** — реалистичен ли интент в данной ситуации
""")

    st.markdown("### Архитектура")
    st.markdown("""
```
                                    ┌─────────────────┐
                                    │  Эксперт 1      │
┌──────────┐   ┌─────────────┐  ┌──│  (t=0.2)        │──┐
│ XLSX     │──▶│ Парсинг     │──┤  └─────────────────┘  │  Согласны?
│ файлы    │   │ + метаданные│  │  ┌─────────────────┐  ├──── Да ──▶ Результат
└──────────┘   └─────────────┘  └──│  Эксперт 2      │──┤
                                   │  (t=0.8)        │  │
                                   └─────────────────┘  └── Нет ──▶ Судья (t=0.1)
                                                                         │
                                                                         ▼
                                                                    Результат
```
Результат записывается в CSV **инкрементально** (каждая реплика сразу).
Поддерживается `--resume` для продолжения прерванного прогона.
""")

    st.markdown("### Системный промпт оценщика")
    st.caption("Один вызов API = оценка по всем 3 критериям")

    # Import actual prompt from run.py
    try:
        from eval_intents.run import EVALUATOR_SYSTEM
        st.code(EVALUATOR_SYSTEM, language="text")
    except ImportError:
        st.warning("Не удалось загрузить промпт из eval_intents/run.py")

    st.markdown("### Системный промпт судьи")
    try:
        from eval_intents.run import JUDGE_SYSTEM
        st.code(JUDGE_SYSTEM, language="text")
    except ImportError:
        st.warning("Не удалось загрузить промпт судьи")

    st.markdown("### Промпт пользователя (для каждой реплики)")
    try:
        from eval_intents.run import EVAL_PROMPT
        st.code(EVAL_PROMPT, language="text")
    except ImportError:
        pass

    st.markdown("""
**Переменные:**
- `{context}` — последние 3 обмена перед оцениваемой репликой
- `{teacher_before}` — реплика репетитора непосредственно перед оцениваемой
- `{student_text}` — текст реплики ученика
- `{intent}` — объявленный интент из тега `[...]`
""")

    st.markdown("### Формат ответа модели")
    st.code(
        '{"is_defect": false, "defect_reason": "",\n'
        ' "match": true, "confidence": 0.95, "reason": "...",\n'
        ' "actual_intent": "",\n'
        ' "ctx_appropriate": true, "ctx_reason": "..."}',
        language="json",
    )

    st.markdown("### Как запустить")
    st.code(
        "# Полный прогон\n"
        "python eval_intents/run.py --input-dir data/intent_eval --output data/intent_eval/results.csv --rpm 100\n\n"
        "# Продолжить прерванный\n"
        "python eval_intents/run.py --input-dir data/intent_eval --output data/intent_eval/results.csv --rpm 100 --resume\n\n"
        "# Тест на 50 репликах\n"
        "python eval_intents/run.py --input-dir data/intent_eval --output data/intent_eval/results.csv --sample 50",
        language="bash",
    )

    st.markdown("### Ручная проверка")
    st.markdown("""
- Результаты модели хранятся в `results.csv`
- Ручные правки хранятся в `human_overrides.json` (отдельно от CSV)
- При загрузке данных правки применяются поверх модельных оценок
- Правки имеют приоритет над моделью по всем 3 критериям: is_defect, match, ctx_appropriate
- Чекбокс «Проверено» отслеживает прогресс ручной валидации
""")

    st.markdown("### Структура results.csv")
    st.dataframe(
        pd.DataFrame({
            "Колонка": [
                "dialog_idx", "turn_idx", "intent_declared", "student_text",
                "teacher_text_before", "is_defect", "defect_reason",
                "match", "confidence", "reason", "actual_intent",
                "ctx_appropriate", "ctx_reason",
                "grade_group", "task_id", "model", "student_type",
                "eval_1", "eval_2", "judge_reasoning",
            ],
            "Тип": [
                "int", "int", "str", "str",
                "str", "bool", "str",
                "bool", "float", "str", "str",
                "bool", "str",
                "str", "str", "str", "str",
                "json", "json", "str",
            ],
            "Описание": [
                "Индекс диалога", "Номер реплики ученика",
                "Интент из тега [...]", "Текст реплики ученика",
                "Текст репетитора", "Бракованная реплика", "Причина брака",
                "Совпал ли интент", "Уверенность (0-1)", "Обоснование", "Фактический интент",
                "Уместен ли интент в контексте", "Обоснование уместности",
                "Класс", "ID задачи", "Модель генерации", "Тип ученика",
                "JSON оценки эксперта 1 (t=0.2)", "JSON оценки эксперта 2 (t=0.8)", "Обоснование судьи",
            ],
        }),
        use_container_width=True,
        hide_index=True,
    )
