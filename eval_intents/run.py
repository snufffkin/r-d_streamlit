#!/usr/bin/env python3
"""
Pipeline for evaluating intent compliance of synthetic student turns.

Uses 3 evaluators at different temperatures + a judge for final verdict.
Writes results incrementally to CSV with full decision log.

Usage:
    python run.py --input-dir data/intent_eval --output data/intent_eval/results.csv
    python run.py --input-dir data/intent_eval --output results.csv --sample 50
    python run.py --input-dir data/intent_eval --output results.csv --resume
"""

import argparse
import asyncio
import csv
import json
import os
import random
import re
import sys
import threading
import time
from dataclasses import dataclass, asdict, field
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

import openpyxl
from openai import AsyncOpenAI
from tqdm import tqdm


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class UserTurn:
    dialog_idx: int
    turn_idx: int
    intent_declared: str
    student_text: str
    teacher_text_before: str
    grade_group: str
    task_id: str
    model: str = ""
    student_type: str = ""


@dataclass
class EvalResult:
    dialog_idx: int
    turn_idx: int
    intent_declared: str
    student_text: str
    teacher_text_before: str
    is_defect: bool = False
    defect_reason: str = ""
    match: bool = True
    confidence: float = 0.0
    reason: str = ""
    actual_intent: str = ""
    ctx_appropriate: bool = True
    ctx_reason: str = ""
    grade_group: str = ""
    task_id: str = ""
    model: str = ""
    student_type: str = ""
    eval_1: str = ""
    eval_2: str = ""
    judge_reasoning: str = ""


CSV_FIELDNAMES = [
    "dialog_idx", "turn_idx", "intent_declared", "student_text",
    "teacher_text_before", "is_defect", "defect_reason",
    "match", "confidence", "reason", "actual_intent",
    "ctx_appropriate", "ctx_reason",
    "grade_group", "task_id", "model", "student_type",
    "eval_1", "eval_2", "judge_reasoning",
]

EVAL_TEMPERATURES = [0.2, 0.8]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_dialog(dialog_text: str) -> list[dict]:
    turns: list[dict] = []
    lines = dialog_text.strip().split("\n")
    current_role = None
    current_intent = None
    current_text_lines: list[str] = []

    def flush():
        if current_role and current_text_lines:
            turns.append({
                "role": current_role,
                "intent": current_intent,
                "text": "\n".join(current_text_lines).strip(),
            })

    for line in lines:
        m = re.match(r"^Пользователь\s*\[([^\]]+)\]\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "user"
            current_intent = m.group(1).strip()
            current_text_lines = [m.group(2).strip()] if m.group(2).strip() else []
            continue
        m = re.match(r"^Пользователь\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "user"
            current_intent = None
            current_text_lines = [m.group(1).strip()] if m.group(1).strip() else []
            continue
        m = re.match(r"^Ассистент\s*:\s*(.*)$", line)
        if m:
            flush()
            current_role = "assistant"
            current_intent = None
            current_text_lines = [m.group(1).strip()] if m.group(1).strip() else []
            continue
        if current_role:
            current_text_lines.append(line)

    flush()
    return turns


def extract_user_turns(
    dialog_text: str, dialog_idx: int, grade_group: str, task_id: str,
    model: str = "", student_type: str = "",
) -> list[UserTurn]:
    turns = parse_dialog(dialog_text)
    result: list[UserTurn] = []
    turn_counter = 0
    for i, turn in enumerate(turns):
        if turn["role"] != "user":
            continue
        turn_counter += 1
        if turn["intent"] is None:
            continue
        teacher_before = ""
        for j in range(i - 1, -1, -1):
            if turns[j]["role"] == "assistant":
                teacher_before = turns[j]["text"]
                break
        result.append(UserTurn(
            dialog_idx=dialog_idx, turn_idx=turn_counter,
            intent_declared=turn["intent"], student_text=turn["text"],
            teacher_text_before=teacher_before,
            grade_group=grade_group, task_id=task_id,
            model=model, student_type=student_type,
        ))
    return result


def build_context_window(dialog_text: str, target_turn_idx: int) -> str:
    turns = parse_dialog(dialog_text)
    user_count = 0
    target_pos = -1
    for i, t in enumerate(turns):
        if t["role"] == "user":
            user_count += 1
            if t["intent"] is not None and user_count == target_turn_idx:
                target_pos = i
                break
    if target_pos < 0:
        return ""
    context_turns = turns[max(0, target_pos - 6):target_pos]
    lines = []
    for t in context_turns:
        role = "Ученик" if t["role"] == "user" else "Учитель"
        intent_tag = f" [{t['intent']}]" if t.get("intent") else ""
        lines.append(f"{role}{intent_tag}: {t['text']}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Filename metadata
# ---------------------------------------------------------------------------

STUDENT_TYPE_MAP = {"medium": "medium", "otlichnik": "strong", "weak": "weak"}
MODEL_MAP = {
    "deepseekV31Terminus": "DeepSeek-V3.1",
    "gemini25falsh": "Gemini-2.5-Flash",
    "gemini3flash": "Gemini-3-Flash",
    "glm45": "GLM-4.5",
}


def parse_filename_metadata(path: str) -> tuple[str, str]:
    stem = Path(path).stem
    for raw_student, norm_student in STUDENT_TYPE_MAP.items():
        for raw_model, norm_model in MODEL_MAP.items():
            if f"_{raw_student}_{raw_model}" in stem:
                return norm_model, norm_student
    return "unknown", "unknown"


# ---------------------------------------------------------------------------
# XLSX reading
# ---------------------------------------------------------------------------

def read_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if len(row) < 19:
            continue
        dialog_text = row[18]
        if not dialog_text or not str(dialog_text).strip():
            continue
        if str(dialog_text).strip() in ("string", "any", "int64", "float64"):
            continue
        records.append({
            "dialog_idx": row_idx - 2,
            "grade_group": str(row[2] or ""),
            "task_id": str(row[8] or ""),
            "dialog": str(dialog_text),
        })
    wb.close()
    return records


# ---------------------------------------------------------------------------
# Rate limiter
# ---------------------------------------------------------------------------

class TokenBucketLimiter:
    def __init__(self, rpm: int):
        self.rpm = rpm
        self.tokens = float(rpm)
        self._last_refill = time.monotonic()
        self._lock = asyncio.Lock()

    async def acquire(self):
        async with self._lock:
            now = time.monotonic()
            self.tokens = min(float(self.rpm), self.tokens + (now - self._last_refill) * self.rpm / 60.0)
            self._last_refill = now
            if self.tokens < 1:
                wait = (1 - self.tokens) * 60.0 / self.rpm
                await asyncio.sleep(wait)
                self.tokens = 0
                self._last_refill = time.monotonic()
            else:
                self.tokens -= 1


# ---------------------------------------------------------------------------
# Incremental CSV writer (thread-safe)
# ---------------------------------------------------------------------------

class IncrementalCSVWriter:
    def __init__(self, path: str, fieldnames: list[str], resume: bool = False):
        self.path = path
        self.fieldnames = fieldnames
        self._lock = threading.Lock()
        self._existing_keys: set[str] = set()

        if resume and Path(path).exists():
            self._load_existing()
            self._file = open(path, "a", newline="", encoding="utf-8")
            self._writer = csv.DictWriter(self._file, fieldnames=fieldnames)
        else:
            self._file = open(path, "w", newline="", encoding="utf-8")
            self._writer = csv.DictWriter(self._file, fieldnames=fieldnames)
            self._writer.writeheader()
            self._file.flush()

    def _load_existing(self):
        with open(self.path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                self._existing_keys.add(self._row_key(row))

    @staticmethod
    def _row_key(row: dict) -> str:
        return f"{row.get('model','')}|{row.get('student_type','')}|{row.get('dialog_idx','')}|{row.get('turn_idx','')}"

    def is_done(self, turn: UserTurn) -> bool:
        return f"{turn.model}|{turn.student_type}|{turn.dialog_idx}|{turn.turn_idx}" in self._existing_keys

    def write(self, result: EvalResult):
        with self._lock:
            self._writer.writerow(asdict(result))
            self._file.flush()

    @property
    def n_existing(self) -> int:
        return len(self._existing_keys)

    def close(self):
        self._file.close()


# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

EVALUATOR_SYSTEM = """\
Ты — эксперт по оценке качества синтетических диалогов «ученик-репетитор» по математике.

Оцени реплику ученика по ТРЁМ критериям последовательно:

═══ ШАГ 1: is_defect — проверка на брак ═══

Реплика бракованная если:
- Текст ЦЕЛИКОМ на английском (несколько полных предложений по-английски). \
НЕ брак: отдельные английские слова/заимствования в русском тексте ("окей", "ок", "лол", "x", "FOIL", "hello"). \
Короткие разговорные реплики типа "окей", "ок", "yes", "no" — это НЕ брак.
- Протекание промпта/инструкций модели (видны системные указания, мета-комментарии, теги вроде [thinking], <system>)
- Отказ модели: «Извините, но я не могу выполнить этот запрос», «I cannot», «As an AI» и т.п.
- Текст слишком длинный для реплики школьника (больше 50 слов)
- Бессмысленный набор символов или повторений

Если is_defect=true — дальнейшая оценка не нужна, match и ctx_appropriate можно не заполнять.

═══ ШАГ 2: match — соответствует ли текст реплики объявленному интенту? ═══

Интенты:
- **answer** — ученик ПЫТАЕТСЯ ответить. Любая попытка дать ответ: число, шаг решения, рассуждение. \
КРИТИЧЕСКИ ВАЖНО: неправильный ответ — это ВСЁ РАВНО answer. Ты оцениваешь НАМЕРЕНИЕ ответить, а НЕ правильность. \
«7*8=54» — answer (неверный, но пытается). «корень из 9 это 4» — answer. «x равен 3» когда правильно 5 — answer. \
НЕ answer только если ученик вообще не пытается отвечать (задаёт вопрос, просит, соглашается).
- **get-explanation** — НЕ решает. Просит объяснить что-то из реплики репетитора.
- **get-solution** — просит дать готовый ответ / решить за него.
- **agree-with-tutor** — короткое подтверждение (1-2 слова): «понял», «ок». НЕ ответ, НЕ вопрос.
- **chat** — уводит от учёбы, не по теме.
- **thank-tutor** — благодарит за конкретное действие.
- **set-problem** — игнорирует вопрос, предлагает свою задачу.
- **criticize-tutor** — недоволен репликой репетитора.

═══ ШАГ 3: ctx_appropriate — реалистичен ли этот интент в данной ситуации? ═══

Это НЕ проверка соответствия текста интенту (это шаг 2). \
Это проверка: мог ли реальный ученик в этой ситуации отреагировать таким интентом?

Вопрос: «Реалистично ли, что ученик в ответ на ЭТУ реплику репетитора выберет интент [{intent}]?»

Неуместно (ctx_appropriate=false):
- Репетитор задал конкретный вопрос/задание — а интент agree-with-tutor или thank-tutor (нереалистично игнорировать вопрос)
- Репетитор просто объясняет (нет вопроса) — а интент answer (нечего решать)
- Репетитор спрашивает «понятно?/готов?» — а интент answer (это не задание)

Уместно (ctx_appropriate=true):
- Репетитор задал вопрос — answer, get-explanation, get-solution
- Репетитор объяснил — agree-with-tutor, get-explanation, thank-tutor
- chat, set-problem, criticize-tutor — поведенческие, реалистичны в любой момент

JSON: {"is_defect": true/false, "defect_reason": "если брак", "match": true/false, "confidence": 0.0-1.0, "reason": "...", "actual_intent": "если отличается", "ctx_appropriate": true/false, "ctx_reason": "..."}"""

EVAL_PROMPT = """\
Контекст: {context}
Реплика репетитора: {teacher_before}
Реплика ученика: {student_text}
Объявленный интент: [{intent}]"""

JUDGE_SYSTEM = """\
Ты — судья. Тебе даны 2 оценки одной реплики. Каждая содержит is_defect, match, ctx_appropriate.

Вынеси финальный вердикт по ВСЕМ критериям, выбрав наиболее убедительные аргументы.

JSON: {"is_defect": true/false, "defect_reason": "...", "match": true/false, "confidence": 0.0-1.0, "reason": "...", "actual_intent": "...", "ctx_appropriate": true/false, "ctx_reason": "...", "reasoning": "анализ 2 оценок"}"""

JUDGE_PROMPT = """\
Реплика репетитора: {teacher_before}
Реплика ученика: {student_text}
Объявленный интент: [{intent}]

Оценка эксперта 1 (t=0.2): {eval_1}
Оценка эксперта 2 (t=0.8): {eval_2}

Вынеси финальный вердикт."""


# ---------------------------------------------------------------------------
# API calls
# ---------------------------------------------------------------------------

async def call_api(
    client: AsyncOpenAI,
    limiter: TokenBucketLimiter,
    eval_model: str,
    system: str,
    prompt: str,
    temperature: float,
    max_retries: int = 5,
) -> dict | None:
    """Single API call with rate limiting and retries. Returns parsed JSON or None."""
    for attempt in range(max_retries):
        await limiter.acquire()
        try:
            response = await client.chat.completions.create(
                model=eval_model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": prompt},
                ],
                temperature=temperature,
                max_tokens=512,
                response_format={"type": "json_object"},
            )
            text = response.choices[0].message.content.strip()
            if text.startswith("```"):
                text = re.sub(r"^```(?:json)?\s*", "", text)
                text = re.sub(r"\s*```$", "", text)
            parsed = json.loads(text)
            if isinstance(parsed, list):
                parsed = parsed[0] if parsed else {}
            return parsed

        except json.JSONDecodeError:
            if attempt < max_retries - 1:
                await asyncio.sleep(1)
                continue
            return None

        except Exception as e:
            if attempt < max_retries - 1:
                wait = min(2 ** (attempt + 2), 60) if "429" in str(e) else 2 ** attempt
                await asyncio.sleep(wait)
                continue
            return None

    return None


async def evaluate_turn_ensemble(
    client: AsyncOpenAI,
    turn: UserTurn,
    dialog_text: str,
    limiter: TokenBucketLimiter,
    eval_model: str,
) -> EvalResult:
    """Run 2 evaluators (t=0.2, t=0.8) + judge on disagreement."""
    context = build_context_window(dialog_text, turn.turn_idx)
    teacher = turn.teacher_text_before or "(нет)"

    prompt = EVAL_PROMPT.format(
        context=context or "(начало диалога)",
        teacher_before=teacher,
        student_text=turn.student_text,
        intent=turn.intent_declared,
    )

    # 2 evaluators in parallel
    eval_tasks = [
        call_api(client, limiter, eval_model, EVALUATOR_SYSTEM, prompt, temp)
        for temp in EVAL_TEMPERATURES
    ]
    eval_results = list(await asyncio.gather(*eval_tasks))

    default = {"is_defect": False, "defect_reason": "", "match": False, "confidence": 0.0,
               "reason": "API error", "actual_intent": "", "ctx_appropriate": True, "ctx_reason": ""}
    eval_strs = []
    for i, r in enumerate(eval_results):
        if r is None:
            eval_results[i] = dict(default)
        eval_strs.append(json.dumps(eval_results[i], ensure_ascii=False))

    def _build_result(src: dict, judge_text: str = "") -> EvalResult:
        return EvalResult(
            dialog_idx=turn.dialog_idx, turn_idx=turn.turn_idx,
            intent_declared=turn.intent_declared,
            student_text=turn.student_text,
            teacher_text_before=turn.teacher_text_before,
            is_defect=bool(src.get("is_defect", False)),
            defect_reason=str(src.get("defect_reason", "")),
            match=bool(src.get("match", False)),
            confidence=float(src.get("confidence", 0.0)),
            reason=str(src.get("reason", "")),
            actual_intent=str(src.get("actual_intent", "")),
            ctx_appropriate=bool(src.get("ctx_appropriate", True)),
            ctx_reason=str(src.get("ctx_reason", "")),
            grade_group=turn.grade_group, task_id=turn.task_id,
            model=turn.model, student_type=turn.student_type,
            eval_1=eval_strs[0], eval_2=eval_strs[1],
            judge_reasoning=judge_text,
        )

    # Check consensus on all 3 dimensions
    defects = [r.get("is_defect", False) for r in eval_results]
    matches = [r.get("match", False) for r in eval_results]
    ctx_votes = [r.get("ctx_appropriate", True) for r in eval_results]

    if len(set(defects)) == 1 and len(set(matches)) == 1 and len(set(ctx_votes)) == 1:
        best = max(eval_results, key=lambda r: r.get("confidence", 0.0))
        return _build_result(best, f"Unanimous (defect={defects[0]}, match={matches[0]}, ctx={ctx_votes[0]})")

    # Disagreement — call judge
    judge_prompt = JUDGE_PROMPT.format(
        teacher_before=teacher,
        student_text=turn.student_text,
        intent=turn.intent_declared,
        eval_1=eval_strs[0], eval_2=eval_strs[1],
    )
    judge_result = await call_api(client, limiter, eval_model, JUDGE_SYSTEM, judge_prompt, 0.1)

    if judge_result is None:
        return _build_result(eval_results[0], "Judge API failed, fallback to eval_1")

    return _build_result(judge_result, str(judge_result.get("reasoning", "")))


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

async def run_pipeline(
    input_path: str,
    output_path: str,
    sample_n: int | None = None,
    rpm: int = 200,
    input_dir: str | None = None,
    resume: bool = False,
    eval_model: str = "google/gemini-3-flash-preview",
):
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        print("ERROR: OPENROUTER_API_KEY environment variable is not set.")
        sys.exit(1)

    client = AsyncOpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )

    # Collect input files
    if input_dir:
        xlsx_files = sorted(Path(input_dir).glob("_rendered*.xlsx"))
        if not xlsx_files:
            print(f"ERROR: No _rendered*.xlsx files found in {input_dir}")
            sys.exit(1)
        print(f"Found {len(xlsx_files)} XLSX files in {input_dir}")
    else:
        xlsx_files = [Path(input_path)]

    # Read all XLSX files
    all_turns: list[tuple[UserTurn, str]] = []
    for xlsx_path in xlsx_files:
        model_name, student_type = parse_filename_metadata(str(xlsx_path))
        print(f"Reading {xlsx_path.name} (model={model_name}, student={student_type})...")
        records = read_xlsx(str(xlsx_path))
        print(f"  Found {len(records)} dialogs.")
        for rec in records:
            turns = extract_user_turns(
                rec["dialog"], rec["dialog_idx"], rec["grade_group"], rec["task_id"],
                model=model_name, student_type=student_type,
            )
            for t in turns:
                all_turns.append((t, rec["dialog"]))

    print(f"\nTotal: {len(all_turns)} user turns with intent tags.")

    # Shuffle so turns from different files are interleaved
    random.seed(42)
    random.shuffle(all_turns)

    if sample_n is not None and sample_n < len(all_turns):
        all_turns = all_turns[:sample_n]
        print(f"  Sampled {sample_n} turns.")

    # Setup CSV writer with resume
    csv_writer = IncrementalCSVWriter(output_path, CSV_FIELDNAMES, resume=resume)
    if resume and csv_writer.n_existing > 0:
        before = len(all_turns)
        all_turns = [(t, d) for t, d in all_turns if not csv_writer.is_done(t)]
        print(f"  Resume: {csv_writer.n_existing} done, {len(all_turns)} remaining")

    if not all_turns:
        print("Nothing to evaluate — all done.")
        csv_writer.close()
        return

    # Each turn = 2 evals + ~15% judge = ~2.3 avg
    est_calls = len(all_turns) * 2.3
    est_min = est_calls / rpm
    print(f"\nEvaluating {len(all_turns)} turns via {eval_model}")
    print(f"  3 checks: defect + intent match + context appropriateness")
    print(f"  2 evaluators (t=0.2, 0.8) + judge on disagreement")
    print(f"  ~{est_calls:.0f} API calls at {rpm} RPM = ~{est_min:.0f} min ({est_min/60:.1f} hours)")

    # Concurrency = rpm/2 (each turn = 2 API calls)
    max_concurrent = max(5, rpm // 2)
    limiter = TokenBucketLimiter(rpm=rpm)
    semaphore = asyncio.Semaphore(max_concurrent)
    pbar = tqdm(total=len(all_turns), desc="Evaluating", unit="turn")
    stats = {"matched": 0, "total": 0, "errors": 0, "ctx_inappropriate": 0, "defects": 0}
    print(f"  Concurrency: {max_concurrent} parallel turns")

    async def process(turn: UserTurn, dialog: str):
        async with semaphore:
            result = await evaluate_turn_ensemble(client, turn, dialog, limiter, eval_model)
        csv_writer.write(result)
        stats["total"] += 1
        if result.is_defect:
            stats["defects"] += 1
        if result.match:
            stats["matched"] += 1
        if not result.ctx_appropriate:
            stats["ctx_inappropriate"] += 1
        if "error" in result.reason.lower():
            stats["errors"] += 1
        pbar.update(1)

    tasks = [process(turn, dialog) for turn, dialog in all_turns]
    await asyncio.gather(*tasks)
    pbar.close()
    csv_writer.close()

    total = stats["total"]
    matched = stats["matched"]
    ctx_bad = stats["ctx_inappropriate"]
    defects = stats["defects"]
    print(f"\nDone! {output_path}")
    print(f"  Defects: {defects}/{total} ({100*defects/total:.1f}%)")
    print(f"  Intent match: {matched}/{total} ({100*matched/total:.1f}%)")
    print(f"  Context inappropriate: {ctx_bad}/{total} ({100*ctx_bad/total:.1f}%)")
    if stats["errors"]:
        print(f"  Errors: {stats['errors']}")


def main():
    parser = argparse.ArgumentParser(description="Evaluate intent compliance via Gemini (ensemble + judge).")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input", help="Single XLSX file.")
    group.add_argument("--input-dir", help="Directory with _rendered*.xlsx files.")
    parser.add_argument("--output", required=True, help="Output CSV path.")
    parser.add_argument("--sample", type=int, default=None)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--model", default="google/gemini-3-flash-preview")
    parser.add_argument("--rpm", type=int, default=200, help="Requests per minute (default: 200).")
    args = parser.parse_args()

    asyncio.run(run_pipeline(
        input_path=args.input or "",
        output_path=args.output,
        sample_n=args.sample,
        rpm=args.rpm,
        input_dir=args.input_dir,
        resume=args.resume,
        eval_model=args.model,
    ))


if __name__ == "__main__":
    main()
